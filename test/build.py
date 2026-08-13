"""
build.py — China PD Monthly Report HTML builder

Reads three xlsx data sources, joins by SKU (exact match only — never fuzzy),
and renders the report HTML by substituting JSON into template.html.

Sources:
  1. Weekly Tracker/China_PD_Weekly_Tracker_WK{xx}.xlsx  (project progress, ground truth)
  2. Monthly PD Report/Summers_Monthly_PD_Table.xlsx     (commercial info)
  3. Monthly PD Report/Project list.xlsx — China Projects sheet  (Sales white-list)

Output:
  Monthly PD Report/China_PD_Monthly_Report_{Mon}{Year}.html

Rotation rule:
  Within the same month-filename, keep latest + previous (_prev suffix).
  Older versions are deleted on each run.
"""

import base64
import io
import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl as ox

# Windows console defaults to cp1252 and chokes on Chinese/emoji output.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, 'reconfigure'):
        _stream.reconfigure(encoding='utf-8', errors='replace')

# -------------------------------------------------------------
# Paths — derive from __file__ so the script is session-id independent.
# This file lives at <BASE>/Monthly PD Report/build.py, so BASE is two levels up.
# -------------------------------------------------------------
MONTHLY_DIR = Path(__file__).resolve().parent
# TEST COPY (2026-08-10): lives in <original>/test/, one level deeper than the
# real build.py, so BASE needs one extra .parent to keep pointing at the PMO
# folder (Weekly Tracker lookup). Everything else resolves inside test/.
BASE = MONTHLY_DIR.parent.parent
WEEKLY_DIR = BASE / 'Weekly Tracker'
# Sandbox uploads dir (Cowork-injected). Fallback when OneDrive Files On-Demand
# corrupts the project copy. Check both for the freshest readable Tracker.
# Derive from $HOME (Cowork sets HOME=/sessions/<id>) so it follows the current
# session automatically; the prior hardcoded session id breaks on every reboot.
_HOME = os.environ.get('HOME', '')
if _HOME and '/sessions/' in _HOME:
    UPLOADS_DIR = Path(_HOME) / 'mnt' / 'uploads'
else:
    UPLOADS_DIR = Path('/nonexistent-uploads')


def _safe_exists(p):
    """Path.exists() that tolerates PermissionError on cross-session mounts."""
    try:
        return p.exists()
    except (PermissionError, OSError):
        return False


def _find_latest_tracker():
    """Highest WKn xlsx that openpyxl can actually open. Falls back to uploads
    if the OneDrive copy has a corrupt zip footer."""
    candidates = []
    for d in (WEEKLY_DIR, UPLOADS_DIR):
        if _safe_exists(d):
            candidates.extend(d.glob('China_PD_Weekly_Tracker_WK*.xlsx'))
            candidates.extend(d.glob('China PD Weekly Tracker WK*.xlsx'))
    candidates = [p for p in candidates if 'backup' not in p.name.lower()]
    if not candidates:
        return WEEKLY_DIR / 'China_PD_Weekly_Tracker_WK17.xlsx'  # legacy default
    def wk_num(p):
        import re as _re
        m = _re.search(r'WK(\d+)', p.name)
        return int(m.group(1)) if m else 0
    candidates.sort(key=lambda p: (wk_num(p), p.stat().st_mtime), reverse=True)
    for p in candidates:
        try:
            ox.load_workbook(p, data_only=True)
            return p
        except Exception:
            continue
    return candidates[0]  # let main() raise the real error


TRACKER_PATH = _find_latest_tracker()


def _find_prev_tracker(current_path):
    """WK 号次大的 tracker (Weekly Tracker 主目录 + Archive)。CRD Change 自动 diff 的基线。
    2026-07-07 加: CRD Change tab 从 review-gated 手填改为每次 build 自动对比上一周。"""
    import re as _re
    def wk_num(p):
        m = _re.search(r'WK(\d+)', p.name)
        return int(m.group(1)) if m else 0
    cur_wk = wk_num(current_path)
    candidates = []
    for d in (WEEKLY_DIR, WEEKLY_DIR / 'Archive'):
        if _safe_exists(d):
            candidates.extend(d.glob('China_PD_Weekly_Tracker_WK*.xlsx'))
    candidates = [p for p in candidates if 'backup' not in p.name.lower() and 0 < wk_num(p) < cur_wk]
    if not candidates:
        return None
    candidates.sort(key=wk_num, reverse=True)
    for p in candidates:
        try:
            ox.load_workbook(p, data_only=True)
            return p
        except Exception:
            continue
    return None


def _crd_date(s):
    """CRD 单元格文本 → date。只认无歧义写法 (YYYY-MM-DD 或 M/D/YY[YY])，纯文字/无年份的跳过。"""
    import re as _re
    from datetime import date as _date
    if not s:
        return None
    s = str(s).strip()
    m = _re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        try:
            return _date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = _re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if m:
        mo, da, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        yr = 2000 + yr if yr < 100 else yr
        try:
            return _date(yr, mo, da)
        except ValueError:
            return None
    return None


# 2026-07-28 Summer 定：CRD 格里的延期字样也算 delay（不要求能解析成日期）
_CRD_DELAY_TXT = re.compile(r'delay|delayed|postpon|延期|推迟|延至|逆期|逾期', re.I)

def _delay_reason(r, kind):
    """Delay Reason 取值：卡点列优先；delay 类若卡点没写延期/失败原因，则并入 Action 里相关的句子。"""
    issue = (r.get('issue') or '').strip()
    nxt = (r.get('next_action') or '').strip()
    if kind != 'delay':
        return issue or nxt
    key = re.compile(r'fail|rework|返工|不合格|delay|延期|推迟|延至', re.I)
    if issue and key.search(issue):
        return issue
    hits = [seg.strip() for seg in re.split(r'\s*/\s*', nxt) if key.search(seg)]
    if hits:
        return (issue + ' / ' if issue else '') + ' / '.join(hits)
    return issue or nxt

def compute_crd_changes(tracker_rows, config, sku_aliases):
    """自动计算 CRD Change / Possible Delay（2026-07-07 起取代 config 手填的 crd_changes）。

    口径（沿用 6-30 Summer 定稿）：
      delay = CRD 比上一周 Tracker 推迟 + 有 PO（提前的不收）
      risk  = 高风险 + 有 PO（CRD 未动，但可能 delay）
    Delay Reason 默认取 Tracker 卡点列；config `crd_change_overrides` 可按 SKU
    覆盖 reason/reasonEN 或 suppress 剔除误报。
    """
    prev_path = _find_prev_tracker(TRACKER_PATH)
    if prev_path is None:
        print('      CRD change: no previous-week tracker found -> empty')
        return []
    # 2026-07-28 修：原来用别名后的 sku 做 key，-CA 变体和母行会被折叠成一条（dict 后者覆盖前者），
    # 导致母行的真实延期丢失、tab 里显示的却是 -CA 行的数据挂在母行名下。
    # 例：RJ11-18-SCTI-HP-V2 7/20→8/2 延期消失，只剩 -CA 行的 8/28 at-risk。
    # 改用**别名前的原始 SKU**逐行配对（load_tracker 已存 sku_raw）。
    _rawkey = lambda row: normalize_sku(row.get('sku_raw') or '') or row['sku']
    prev_by_sku = {_rawkey(r): r for r in load_tracker(prev_path, sku_aliases)}
    overrides = {o.get('sku'): o for o in config.get('crd_change_overrides', [])}
    out = []
    for r in tracker_rows:
        po_stat, _ = parse_po(r['po_status'])
        if po_stat != 'YES':
            continue
        ov = overrides.get(r['sku'], {})
        if ov.get('suppress'):
            continue
        prev = prev_by_sku.get(_rawkey(r))
        cur_d = _crd_date(r['crd'])
        prev_d = _crd_date(prev['crd']) if prev else None
        if prev_d and cur_d and cur_d > prev_d:
            kind, days = 'delay', (cur_d - prev_d).days
        elif cur_d is None and _CRD_DELAY_TXT.search(str(r['crd'] or '')):
            # 2026-07-28 Summer 定：CRD 格里写了 delay / 延期 / 推迟 的，
            # 不管能不能解析成日期，一律进 CRD Change tab（例：RJ15-7-LL-DR 验货 fail 后 CRD 直接写 'delay'）
            kind, days = 'delay', 0
        elif '高' in (r['risk'] or ''):
            kind, days = 'risk', 0
        else:
            continue
        _, buyers = parse_po_buyers(r['po_status'])
        buyer_disp = '/'.join(b['buyer'] for b in buyers if b['status'] == 'YES')
        if not buyer_disp:
            buyer_disp = parse_po(r['po_status'])[1]
        fmt = lambda d: d.strftime('%m/%d/%Y') if d else (r['crd'] or '')
        # 文本型延期（cur 无日期）：old 用上周值（日期或原文），new 用当前文本
        _prev_disp = (prev_d.strftime('%m/%d/%Y') if prev_d else ((prev or {}).get('crd') or ''))
        _cur_disp = cur_d.strftime('%m/%d/%Y') if cur_d else str(r['crd'] or '')
        out.append({
            'sku': _rawkey(r), 'pm': r['pm'], 'status': r['current_status'],
            'risk': r['risk'] or '', 'kind': kind,
            'oldCRD': _prev_disp if kind == 'delay' else _cur_disp,
            'newCRD': _cur_disp, 'days': days, 'poBuyer': buyer_disp,
            # 2026-07-28：Delay Reason 只取卡点列会漏掉写在「下一步 Action」里的真实原因
            # （WK30 RJ15-7-LL-DR 的「一轮验货 fail、需 rework」就写在 Action 里）。
            # delay 类条目：卡点没提到 fail/rework/延期 时，把 Action 里相关那句并进来。
            'reason': ov.get('reason') or _delay_reason(r, kind),
            'reasonEN': ov.get('reasonEN', ''),
        })
    out.sort(key=lambda e: (0 if e['kind'] == 'delay' else 1, -e['days']))
    print(f'      CRD change (auto vs {prev_path.name}): {len(out)} entries')
    for e in out:
        arrow = f"{e['oldCRD']} -> {e['newCRD']} (+{e['days']}d)" if e['kind'] == 'delay' else f"{e['newCRD']} at-risk"
        print(f"        [{e['kind']}] {e['sku']} {arrow} ({e['pm']})")
    return out
PDTABLE_PATH = MONTHLY_DIR / 'Summers_Monthly_PD_Table.xlsx'

# Project List retired 2026-07-02: the HTML "For Sales / All" filter was removed,
# so the white-list xlsx is no longer a data source.
# Source of embedded product images. Auto-detect the latest 'China PD updates *.xlsx'
# in MONTHLY_DIR by mtime, so Summer can drop in next month's file (e.g. 'China PD
# updates May 2026.xlsx') without editing build.py.
def _find_latest_pd_updates():
    """Newest readable 'China PD updates *.xlsx'. Project first, then uploads
    (OneDrive Files On-Demand bug)."""
    candidates = []
    for d in (MONTHLY_DIR, UPLOADS_DIR):
        if _safe_exists(d):
            candidates.extend(d.glob('China PD updates *.xlsx'))
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for p in candidates:
        try:
            ox.load_workbook(p, data_only=True)
            return p
        except Exception:
            continue
    return candidates[0] if candidates else None

PDUPDATES_PATH = _find_latest_pd_updates()
TEMPLATE_PATH = MONTHLY_DIR / 'template.html'
TRANSLATIONS_PATH = MONTHLY_DIR / 'translations.json'

# Output naming: month abbreviation + 4-digit year
NOW = datetime.now()
# Report period auto-detection (Summer rule 2026-06-02): the report rolls to
# month M on the 10th of M.  day >= 10 -> current month; day < 10 -> previous
# month (crossing the year boundary in January).
import calendar as _calendar
if NOW.day >= 10:
    _pm, _py = NOW.month, NOW.year
elif NOW.month == 1:
    _pm, _py = 12, NOW.year - 1
else:
    _pm, _py = NOW.month - 1, NOW.year
MONTH_NAME = _calendar.month_abbr[_pm]   # 'May','Jun','Jul' (filenames)
MONTH_FULL = _calendar.month_name[_pm]   # 'May','June','July' (display title)
YEAR = str(_py)
REPORT_PERIOD = f'{MONTH_FULL} {YEAR}'   # e.g. 'June 2026'

# Data cutoff (Summer rule 2026-06-09): data validity is generally through
# last Friday. Compute the most recent Friday on/before the build date.
from datetime import timedelta as _timedelta
_asof = NOW - _timedelta(days=(NOW.weekday() - 4) % 7)   # weekday: Fri=4
DATA_ASOF_CN = f'数据截至 {_asof.year}年{_asof.month}月{_asof.day}日（上周五）'
DATA_ASOF_EN = f'Data as of {_calendar.month_abbr[_asof.month]} {_asof.day}, {_asof.year}'

OUT_NAME = f'China_PD_Monthly_Report_{MONTH_NAME}{YEAR}.html'
OUT_PATH = MONTHLY_DIR / OUT_NAME
PREV_PATH = MONTHLY_DIR / f'China_PD_Monthly_Report_{MONTH_NAME}{YEAR}_prev.html'

# English-translated output (for US Sales)
OUT_NAME_EN = f'China_PD_Monthly_Report_{MONTH_NAME}{YEAR}_EN.html'
OUT_PATH_EN = MONTHLY_DIR / OUT_NAME_EN
PREV_PATH_EN = MONTHLY_DIR / f'China_PD_Monthly_Report_{MONTH_NAME}{YEAR}_EN_prev.html'

# Scratchpad for safe-write (avoids OneDrive zip-truncation issue).
# Use the env var so it follows the current session automatically; outside the
# sandbox (e.g. local Windows) fall back to the OS temp dir.
import tempfile as _tempfile
SCRATCH = Path(os.environ.get('CLAUDE_SCRATCH') or Path(_tempfile.gettempdir()) / 'pd_report_scratch')
SCRATCH.mkdir(parents=True, exist_ok=True)

# -------------------------------------------------------------
# Constants
# -------------------------------------------------------------
PIPELINE_LABELS = [
    'Kick off', 'Detail Design', 'Prototype', 'Tooling',
    'FOT', 'EB', 'Culinary EB', 'Culinary Claims',
    'PP', 'Culinary PP', 'MP',
]
# Inspection merged into MP (Summer 2026-05-04: 'inspection 和 MP 是一样的')

# Map currentStatus value → pipeline label index
STATUS_TO_PIPELINE = {
    'Kick off': 0, 'Kick Off': 0, 'kick off': 0,
    'Detail Design': 1, 'Design': 1,
    'Prototype': 2,
    'Tooling': 3, 'Tooling Launch': 3,
    'FOT': 4,
    'EB': 5, 'EB1': 5, 'EB2': 5,
    'Culinary EB': 6,
    'Culinary Claims': 7,
    'PP': 8,
    'Culinary PP': 9,
    'MP': 10, 'MP中': 10, 'Inspection': 10,
}

# PM section ordering (HTML displays in this order)
PM_SECTION_ORDER = [
    'Cottee Wei — 空气炸锅 + T1 项目',
    'Rowling Luo — 烤箱 / 面包机 / 饭煲 / 慢炖锅 / 油炸锅',
    'Serena Sun — ICEMAN / 咖啡 / 冰淇淋',
    'Chris Zhou — 烤盘 / 搅拌类 + MX 项目',
    'Liz Liu — 水壶 + 微波炉',
    # 7/14: Jenifer 是 2026-06 新增的第 6 位报送人；此前漏加导致她的 PD Table 段
    # （⚠️ Other / 未归类 PM）从不进 Page 1 卡片循环
    'Jenifer Yuan — CSM 杭州 (C60 / C45 / CQ60 / C22)',
]

# -------------------------------------------------------------
# Helpers
# -------------------------------------------------------------
# Canonical category buckets — Summer 2026-05-04. Page 1 cards group by these.
# Order matters: more-specific keywords first (e.g. "Air Fryer (Oven)" must
# match before "Air Fryers" or "Oven"). NEVER add fuzzy/aggressive rules
# without asking Summer first.
CATEGORY_RULES = [
    ('Air Fryer (Oven)', ['air fryer (oven)', 'air fryer oven']),
    ('Air Fryers',       ['air fryer', 'air fry']),
    ('Microwave',        ['microwave']),
    ('Pressure Cooker',  ['pressure']),
    ('Slow Cooker',      ['slow']),
    ('Rice Cooker',      ['rice']),
    ('Deep Fryer',       ['deep fryer', 'deep']),
    ('Roaster Oven',     ['roaster']),
    ('Bread Maker',      ['bread']),
    ('Ice Cream',        ['ice cream']),
    ('Iceman',           ['iceman', 'ice maker', 'icemaker', 'slush']),
    ('Water Dispenser',  ['water dispenser', 'dispenser']),
    ('Coffee',           ['coffee']),
    ('Griddle',          ['griddle']),
    ('Blender',          ['blender']),
    ('Mixer',            ['mixer']),
    ('Vacuum',           ['vacuum']),
    ('Kettle',           ['kettle']),
    ('Oven',             ['oven']),
    ('Grill',            ['grill', 'panini']),
]


def normalize_category(raw):
    """Map a free-text category to one canonical bucket. Empty stays empty;
    unknown strings pass through unchanged so we don't silently lose data."""
    if not raw:
        return ''
    s = str(raw).strip()
    if not s:
        return ''
    sl = s.lower()
    for canonical, kws in CATEGORY_RULES:
        for kw in kws:
            if kw in sl:
                return canonical
    return s


def cellstr(v):
    """Stringify a cell value, treating None as empty."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def normalize_sku(raw):
    """SKU normalization: drop newline annotations, strip whitespace.

    Examples:
      'RJ38-6T-lava grey\\n⚠️待确认项目' → 'RJ38-6T-lava grey'
      'RJ34-10C-D ' → 'RJ34-10C-D'

    NEVER drops semantic suffixes (SS/CA/MX/etc.) — exact match only.
    """
    if not raw:
        return ''
    s = str(raw)
    # Cut at first newline (annotation block)
    s = s.split('\n')[0]
    # 5/12: defensive fullwidth → halfwidth bracket fix
    s = s.replace('（', '(').replace('）', ')')
    return s.strip()


def is_pm_section_header(value):
    """True if cell value is one of the 5 PM section headers."""
    if not value:
        return False
    s = str(value).strip()
    return s in PM_SECTION_ORDER


def is_mx_sku(sku):
    """SKU is part of the Mexico pipeline if its last hyphen-delimited token is MX.
    Examples: 'RJ40-8-MX' -> True, 'RJ55-7-VN-MX / SMR-VN-MX' -> True (ends -MX),
    'RJ50-SFDAF-25D' -> False. Confirmed by Summer 2026-05-04."""
    if not sku:
        return False
    return str(sku).strip().upper().endswith('-MX')


def clean_status(s):
    """Normalize Current Status string to a single-line value."""
    if not s:
        return ''
    return str(s).replace('\n', ' ').strip()


# -------------------------------------------------------------
# Source loaders
# -------------------------------------------------------------
def load_tracker(path, sku_aliases=None):
    """Load Tracker. Returns list of dicts in row-order, plus pm_section list.

    Tracker layout (26 cols, 5/19 added E=NPD/ASI; all later cols shifted +1):
      A=#  B=品类  C=P/V  D=SKU  E=NPD/ASI  F=风险  G=PM  H=Tier  I=上次更新
      J=Current Status  K=卡点/风险  L=下一步Action  M=PA状态  N=PO/订单状态  O=CRD
      P..Z = Kick off … MP (11 stage cols, Inspection removed)

    Each dict has fields:
      num, sku, sku_raw, pv, category, npd_asi, risk, pm, tier, last_update,
      current_status, issue, next_action, po_status, crd,
      stages (dict of stage_label → date_or_check), pm_section

    5/12: sku_aliases (Tracker SKU → canonical PD Table name) applied to `sku`
    after normalize_sku so HTML join uses canonical form. `sku_raw` preserves
    the original Tracker spelling for debug.
    5/19: NPD/ASI col E added; ASI source of truth is now Tracker col E
    instead of pd_table_config.json's after_sales_improvement list.
    """
    if not path.exists():
        raise FileNotFoundError(f'Tracker not found: {path}')

    wb = ox.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    rows = []
    current_section = ''
    aliases = sku_aliases or {}

    # 6/15: resolve columns by HEADER (row 1), not fixed index. A '6A' column was
    # inserted at N (col 14) on 6/15, shifting PO/CRD/stages one column right.
    # Header lookup makes load_tracker robust to future column inserts; falls back
    # to fixed indices when a header is missing (older trackers).
    import re as _re
    def _norm_h(s):
        return _re.sub(r'\s+', ' ', str(s or '').strip()).lower()
    hdr = {}
    for _c in range(1, ws.max_column + 1):
        _h = _norm_h(ws.cell(1, _c).value)
        if _h and _h not in hdr:
            hdr[_h] = _c
    def _col(default, *names):
        for n in names:
            c = hdr.get(_norm_h(n))
            if c:
                return c
        return default
    COL_CAT     = _col(2,  '品类')
    COL_PV      = _col(3,  'P/V')
    COL_SKU     = _col(4,  'SKU')
    COL_NPD     = _col(5,  'NPD/ASI')
    COL_RISK    = _col(6,  '风险')
    COL_PM      = _col(7,  'PM')
    COL_TIER    = _col(8,  'Tier')
    COL_LASTUPD = _col(9,  '上次更新')
    COL_STATUS  = _col(10, 'Current Status')
    COL_ISSUE   = _col(11, '卡点 / 风险', '卡点/风险')
    COL_NEXT    = _col(12, '下一步 Action', '下一步Action')
    COL_PA      = _col(13, 'PA状态')
    COL_6A      = _col(0,  '6A')   # 0 = not present (older trackers)
    COL_PO      = _col(15, 'PO/订单状态')
    COL_CRD     = _col(16, 'CRD')

    # Stage columns (Kick off…MP) removed from Tracker 2026-07-01. HTML never
    # rendered them (front-end has no per-SKU stage timeline; Page 2 buckets by
    # 'Current Status'). Kept empty for downstream compatibility.
    stage_label_map = {}

    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(r, 1).value     # #
        c4 = ws.cell(r, COL_SKU).value     # SKU (header-resolved)
        c1s = cellstr(c1)
        c4s = cellstr(c4)

        # PM section header
        if is_pm_section_header(c1s):
            current_section = c1s
            continue

        # Skip empty/non-data rows
        if not c4s:
            continue

        sku = normalize_sku(c4s)
        if not sku:
            continue
        sku = aliases.get(sku, sku)  # 5/12: apply Tracker→canonical alias

        # Stage cells: '✓' = completed past stage; date = scheduled/done; empty = not yet
        stages = {}
        for col, label in stage_label_map.items():
            v = ws.cell(r, col).value
            stages[label] = cellstr(v)

        rows.append({
            'num': cellstr(c1),
            'sku': sku,
            'sku_raw': c4s,
            'pv': cellstr(ws.cell(r, COL_PV).value),         # 'Parent' / 'Variant' / ''
            'category': cellstr(ws.cell(r, COL_CAT).value),
            'npd_asi': cellstr(ws.cell(r, COL_NPD).value),    # 'ASI' or '' (blank = NPD)
            'risk': cellstr(ws.cell(r, COL_RISK).value),
            'pm': cellstr(ws.cell(r, COL_PM).value),
            'tier': cellstr(ws.cell(r, COL_TIER).value),
            'last_update': cellstr(ws.cell(r, COL_LASTUPD).value),
            'current_status': clean_status(ws.cell(r, COL_STATUS).value),
            'issue': cellstr(ws.cell(r, COL_ISSUE).value),
            'next_action': cellstr(ws.cell(r, COL_NEXT).value),
            'pa_status': ('Done' if cellstr(ws.cell(r, COL_PA).value).strip().lower() == 'yes' else cellstr(ws.cell(r, COL_PA).value)),  # 6/15: report treats non-standard 'Yes' as Done (Summer)
            'six_a': cellstr(ws.cell(r, COL_6A).value) if COL_6A else '',  # 6/15 new col
            'po_status': cellstr(ws.cell(r, COL_PO).value),
            'crd': cellstr(ws.cell(r, COL_CRD).value),
            'stages': stages,
            'pm_section': current_section,
        })
    return rows


def load_pd_table(path):
    """Load Summers Monthly PD Table.

    Returns:
      main_skus: dict {sku → fields}  (rows above '▼ Gap Analysis' marker)
      pending_skus: list of (sku, category, pm_section) for rows in待确认 / Gap sections
    """
    if not path.exists():
        raise FileNotFoundError(f'PD Table not found: {path}')

    wb = ox.load_workbook(path, data_only=True)
    ws = wb['Product Info']

    main_skus = {}
    pending = []
    current_section = ''
    in_pending = False  # flips True at '▼ Gap Analysis' header

    # 24 column mapping (R1 header)
    col_map = {
        1: 'sku', 2: 'category', 3: 'tier', 4: 'brand',
        5: 'description', 6: 'top_feature',
        7: 'uf1', 8: 'uf2', 9: 'uf3',
        10: 'msrp', 11: 'sample_eta', 12: 'po_placed',
        13: 'est_inspection', 14: 'factory', 15: 'market',
        16: 'cost', 17: 'buffer', 18: 'port',
        19: 'duty', 20: 'hc40',
        21: 'comp_model', 22: 'rj_diff',
        23: 'note1', 24: 'note2',
    }

    for r in range(2, ws.max_row + 1):
        c1 = cellstr(ws.cell(r, 1).value)
        c2 = cellstr(ws.cell(r, 2).value)

        # Detect section transitions
        if c1.startswith('▼') or c1.startswith('⚠'):
            in_pending = True
            # Pending sub-section header (e.g., '▼ Cottee — 需补充商业信息')
            continue

        if is_pm_section_header(c1):
            current_section = c1
            in_pending = False
            continue

        if not c1:
            continue

        sku = normalize_sku(c1)
        if not sku:
            continue

        if in_pending:
            pending.append({
                'sku': sku,
                'sku_raw': c1,
                'category': c2,
                'pm_section': current_section,
            })
            continue

        # Main SKU row
        record = {'pm_section': current_section, 'sku_raw': c1}
        for col, key in col_map.items():
            v = ws.cell(r, col).value
            record[key] = cellstr(v)
        # Override sku with normalized
        record['sku'] = sku
        main_skus[sku] = record

    return main_skus, pending


# -------------------------------------------------------------
# Image extraction — pull embedded product renderings out of the PD updates
# xlsx and key them by SKU so we can render <img> tags in the HTML cards.
# -------------------------------------------------------------
# Tunables: thumbnail size (longest side, px) and JPEG quality. 300×~75% gives
# ~30–80 KB per image; 46 images ≈ 2–4 MB total embedded into the HTML.
IMAGE_THUMB_SIZE = 300
IMAGE_JPEG_QUALITY = 78


def _sku_image_aliases(sku):
    """Yield SKU aliases that should share the same product image.

    Confirmed by Summer 2026-04-30: a trailing parenthetical color/material
    code (SS = Stainless Steel, BLK = Black, WHT = White, etc.) names a
    visual variant of the same parent product. The parent SKU (without the
    parenthetical) can therefore reuse the same rendering image.

    Examples:
      'RJ50-SFDAF-25D(SS)'  → 'RJ50-SFDAF-25D(SS)', 'RJ50-SFDAF-25D'
      'RJ50-BFDAF-25D(BLK)' → 'RJ50-BFDAF-25D(BLK)', 'RJ50-BFDAF-25D'
      'RJ38-G4'             → 'RJ38-G4'

    NOTE: This is the ONLY image-aliasing rule. We never prefix-match,
    fuzzy-match, or strip non-parenthetical suffixes — those carry business
    meaning (CA / MX / V2 / etc.) and require explicit Summer approval.
    """
    if not sku:
        return
    yield sku
    m = re.match(r'^(.+?)\(\s*([A-Za-z][A-Za-z0-9]*)\s*\)\s*$', sku)
    if m:
        parent = m.group(1).strip()
        if parent and parent != sku:
            yield parent


def _extract_image_in_cell_raw(path):
    """5/19: Excel 365 'Image in cell' (rich data, NOT image-over-cell) support.

    Image-in-cell stores the image as the cell's *value* via a rich-data
    metadata chain. openpyxl's ws._images doesn't see these. We unzip and
    walk the XML manually:
      cell.vm (1-based) -> metadata.xml/valueMetadata[vm-1].rc.@v
                        -> rdrichvalue.xml/rv[v].<v>0</v> = LocalImageIdentifier L
                        -> richValueRel.xml/rel[L].@r:id = rIdN
                        -> _rels/richValueRel.xml.rels[rIdN] -> media/imageN

    Returns: dict {sheet_name: {(row, col): raw_image_bytes}}
             Empty on missing rich data parts or parse failure.
    """
    import zipfile, re as _re
    from openpyxl.utils import column_index_from_string
    out = {}
    if not path or not path.exists():
        return out
    try:
        with zipfile.ZipFile(path) as z:
            names = set(z.namelist())
            required = ('xl/metadata.xml', 'xl/richData/rdrichvalue.xml',
                        'xl/richData/richValueRel.xml',
                        'xl/richData/_rels/richValueRel.xml.rels',
                        'xl/_rels/workbook.xml.rels', 'xl/workbook.xml')
            if not all(r in names for r in required):
                return out

            # 1) valueMetadata: [(t, v), ...] 0-based
            md_xml = z.read('xl/metadata.xml').decode('utf-8', 'ignore')
            vm_list = []
            vm_block = _re.search(r'<valueMetadata[^>]*>(.*?)</valueMetadata>', md_xml, _re.DOTALL)
            if vm_block:
                for bk in _re.finditer(r'<rc\s+t="(\d+)"\s+v="(\d+)"', vm_block.group(1)):
                    vm_list.append((int(bk.group(1)), int(bk.group(2))))

            # 2) rdrichvalue: rv index -> LocalImageIdentifier (first <v>)
            rv_xml = z.read('xl/richData/rdrichvalue.xml').decode('utf-8', 'ignore')
            rv_local = []
            for m in _re.finditer(r'<rv\b[^>]*>(.*?)</rv>', rv_xml, _re.DOTALL):
                vs = _re.findall(r'<v>([^<]+)</v>', m.group(1))
                rv_local.append(int(vs[0]) if vs else None)

            # 3) richValueRel rel order -> list of rIds
            rvr_xml = z.read('xl/richData/richValueRel.xml').decode('utf-8', 'ignore')
            rvr_rids = _re.findall(r'<rel\s+[^>]*r:id="(rId\d+)"', rvr_xml)

            # 4) rId -> image target
            rels_xml = z.read('xl/richData/_rels/richValueRel.xml.rels').decode('utf-8', 'ignore')
            rid_to_target = {}
            for m in _re.finditer(r'<Relationship\s+Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml):
                rid_to_target[m.group(1)] = m.group(2)

            # 5) sheet name -> worksheet xml path
            wb_xml = z.read('xl/workbook.xml').decode('utf-8', 'ignore')
            wb_rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8', 'ignore')
            rid_to_sheet_target = {}
            for m in _re.finditer(r'<Relationship\s+Id="(rId\d+)"[^>]*Target="(worksheets/[^"]+)"', wb_rels_xml):
                rid_to_sheet_target[m.group(1)] = m.group(2)
            sheet_to_file = {}
            for m in _re.finditer(r'<sheet\s+name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml):
                name = m.group(1).replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
                rid = m.group(2)
                if rid in rid_to_sheet_target:
                    sheet_to_file[name] = 'xl/' + rid_to_sheet_target[rid]

            # 6) Walk each worksheet for cells with vm attribute
            for sheet_name, sheet_file in sheet_to_file.items():
                if sheet_file not in names:
                    continue
                ws_xml = z.read(sheet_file).decode('utf-8', 'ignore')
                for m in _re.finditer(r'<c\s+r="([A-Z]+)(\d+)"[^>]*\svm="(\d+)"', ws_xml):
                    col_letter, row_num, vm = m.group(1), int(m.group(2)), int(m.group(3))
                    col = column_index_from_string(col_letter)
                    if vm <= 0 or vm > len(vm_list):
                        continue
                    t, v = vm_list[vm - 1]
                    if t != 1 or v < 0 or v >= len(rv_local):
                        continue
                    local = rv_local[v]
                    if local is None or local < 0 or local >= len(rvr_rids):
                        continue
                    target = rid_to_target.get(rvr_rids[local])
                    if not target:
                        continue
                    # 'Target' is relative to xl/richData/, e.g. '../media/image8.png'
                    target_path = ('xl/' + target.replace('../', '')).replace('xl/xl/', 'xl/')
                    if target_path not in names:
                        continue
                    out.setdefault(sheet_name, {})[(row_num, col)] = z.read(target_path)
    except Exception as e:
        print(f'      WARN: image-in-cell parse failed ({e}), continuing with image-over-cell only')
        return {}
    return out


def extract_sku_images(path):
    """Extract embedded product images from PD updates xlsx, keyed by SKU.

    Each sheet has products laid out horizontally — col B is the field label,
    cols C+ are one product per column. Row 8 holds the image, Row 10 holds
    the Model (SKU). When a single column carries multiple SKUs (color
    variants stacked in one cell, e.g. RJ64-10-PTC / BTR / LVD / Aqu) and the
    PM has placed several images side-by-side in the same anchor cell, we map
    them by reading order: images sorted top-to-bottom, left-to-right are
    assigned 1:1 to the SKUs in the cell. When there's only one image but
    multiple SKUs, all SKUs share the image. Multi-SKU cells with a single
    image keep the original 'all SKUs share' behaviour.

    Trailing '(SS)' / '(BLK)' / etc. parentheticals are color/material codes —
    we register the bare parent SKU as an additional alias so umbrella PD
    Table entries can match. See _sku_image_aliases().

    5/19: now supports both 'image-over-cell' (traditional, ws._images) and
    'image-in-cell' (Excel 365 rich-data). See _extract_image_in_cell_raw().

    Returns: dict {sku → 'data:image/jpeg;base64,...'} for use as <img src=...>.
             Returns {} on failure (Pillow missing, file unreadable, etc.).
    """
    if not path or not path.exists():
        print(f'      WARN: PD updates file not found, skipping image extraction')
        return {}

    try:
        from PIL import Image as PILImage
    except ImportError:
        print(f'      WARN: Pillow not installed (pip install Pillow '
              '--break-system-packages), skipping image extraction')
        return {}

    images = {}
    skipped = 0
    try:
        wb = ox.load_workbook(path, data_only=True)
    except Exception as e:
        print(f'      WARN: could not open PD updates ({e}), skipping images')
        return {}

    # Bucket size for clustering rowOff values (EMU; 914400 = 1 inch). Images
    # within the same visual row have rowOff differences << 300000 EMU
    # (~0.33"), images in different visual rows differ by > 300000 EMU.
    ROW_BUCKET_EMU = 300000

    def _process_bytes(img_bytes):
        """Process raw image bytes -> base64 JPEG data URI, or '' on error."""
        try:
            pil = PILImage.open(io.BytesIO(img_bytes))
            if pil.mode in ('RGBA', 'LA'):
                bg = PILImage.new('RGB', pil.size, (255, 255, 255))
                bg.paste(pil, mask=pil.split()[-1])
                pil = bg
            elif pil.mode == 'P':
                pil = pil.convert('RGBA')
                bg = PILImage.new('RGB', pil.size, (255, 255, 255))
                bg.paste(pil, mask=pil.split()[-1])
                pil = bg
            elif pil.mode != 'RGB':
                pil = pil.convert('RGB')
            pil.thumbnail((IMAGE_THUMB_SIZE, IMAGE_THUMB_SIZE), PILImage.LANCZOS)
            buf = io.BytesIO()
            pil.save(buf, format='JPEG',
                     quality=IMAGE_JPEG_QUALITY, optimize=True)
            return ('data:image/jpeg;base64,'
                    + base64.b64encode(buf.getvalue()).decode('ascii'))
        except Exception:
            return ''

    def _process(img_obj):
        """Process one openpyxl Image -> base64 JPEG data URI, or '' on error."""
        try:
            img_bytes = img_obj._data()
        except Exception:
            return ''
        try:
            pil = PILImage.open(io.BytesIO(img_bytes))
            if pil.mode in ('RGBA', 'LA'):
                bg = PILImage.new('RGB', pil.size, (255, 255, 255))
                bg.paste(pil, mask=pil.split()[-1])
                pil = bg
            elif pil.mode == 'P':
                pil = pil.convert('RGBA')
                bg = PILImage.new('RGB', pil.size, (255, 255, 255))
                bg.paste(pil, mask=pil.split()[-1])
                pil = bg
            elif pil.mode != 'RGB':
                pil = pil.convert('RGB')
            pil.thumbnail((IMAGE_THUMB_SIZE, IMAGE_THUMB_SIZE), PILImage.LANCZOS)
            buf = io.BytesIO()
            pil.save(buf, format='JPEG',
                     quality=IMAGE_JPEG_QUALITY, optimize=True)
            return ('data:image/jpeg;base64,'
                    + base64.b64encode(buf.getvalue()).decode('ascii'))
        except Exception:
            return ''

    def _register(sku, data_uri):
        for alias in _sku_image_aliases(normalize_sku(sku)):
            if alias and alias not in images:
                images[alias] = data_uri

    # 5/19: Pre-extract image-in-cell raw bytes for all sheets at once
    cell_images = _extract_image_in_cell_raw(path)
    cell_image_count = sum(len(v) for v in cell_images.values())
    if cell_image_count:
        print(f'      image-in-cell: {cell_image_count} cell-anchored images found across {len(cell_images)} sheets')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        has_over = bool(getattr(ws, '_images', None))
        has_in_cell = bool(cell_images.get(sheet_name))
        if not has_over and not has_in_cell:
            continue

        # Build column -> list-of-SKUs map for this sheet from Row 10.
        col_to_skus = {}
        for col_idx in range(3, ws.max_column + 1):  # C onwards
            v = ws.cell(10, col_idx).value
            if not v:
                continue
            skus_in_cell = []
            for line in str(v).split('\n'):
                line = line.strip()
                if not line:
                    continue
                # Strip annotations after the SKU: tab or 2+ spaces both signal
                # 'SKU<sep>colorname' (e.g. 'RJ64-10-PTC\tPistachio' or
                # 'RJ64-10-Aqu  Aqua'). Single space inside an SKU is rare so
                # we only split on tabs and runs of 2+ spaces.
                sku = re.split(r'\t|  +', line, maxsplit=1)[0].strip()
                if sku and any(c.isalnum() for c in sku):
                    skus_in_cell.append(sku)
            if skus_in_cell:
                col_to_skus[col_idx] = skus_in_cell

        # Process image-in-cell first (one image per cell, col matches SKU col)
        if has_in_cell:
            for (row, col), img_bytes in cell_images[sheet_name].items():
                skus = col_to_skus.get(col)
                if not skus:
                    # tolerate 1-col drift like over-cell branch does
                    for offset in (-1, 1):
                        skus = col_to_skus.get(col + offset)
                        if skus:
                            break
                if not skus:
                    skipped += 1
                    continue
                data_uri = _process_bytes(img_bytes)
                if not data_uri:
                    skipped += 1
                    continue
                for sku in skus:
                    _register(sku, data_uri)

        # Then process traditional image-over-cell via openpyxl
        if not has_over:
            continue

        # Group images by their target column (1-indexed).
        # 5/19: skip zero-area "ghost" images (TwoCellAnchor with _from == to).
        # PMs sometimes delete/replace images and Excel keeps the binary in the
        # zip with a collapsed anchor frame (invisible in Excel UI). E.g. an
        # invisible water-bottle PNG in Coffee&Iceman col 5 was getting matched
        # to RJ44-CB instead of its real grinder image.
        images_by_col = {}
        for img in ws._images:
            try:
                fc = img.anchor._from
                tc = getattr(img.anchor, 'to', None)
                col = fc.col + 1
                row_off = fc.rowOff or 0
                col_off = fc.colOff or 0
            except AttributeError:
                skipped += 1
                continue
            # Zero-area filter: TwoCellAnchor where to corner == from corner
            if tc is not None:
                same_col = (tc.col == fc.col and (tc.colOff or 0) == (fc.colOff or 0))
                same_row = (tc.row == fc.row and (tc.rowOff or 0) == (fc.rowOff or 0))
                if same_col or same_row:
                    skipped += 1
                    continue
            images_by_col.setdefault(col, []).append((row_off, col_off, img))

        for col, img_entries in images_by_col.items():
            # Find the SKU cell; tolerate 1-col drift.
            skus = col_to_skus.get(col)
            if not skus:
                for offset in (-1, 1):
                    skus = col_to_skus.get(col + offset)
                    if skus:
                        break
            if not skus:
                skipped += len(img_entries)
                continue

            # Sort by reading order: row bucket, then column offset.
            img_entries.sort(key=lambda t: (t[0] // ROW_BUCKET_EMU, t[1]))

            # When the column contains as many images as SKUs (e.g. ice cream
            # 4 colors, 4 images), assign 1:1 in reading order. Otherwise all
            # SKUs share each image (first one extracted wins per SKU).
            if len(img_entries) >= len(skus) > 1:
                for sku, (_r, _c, img) in zip(skus, img_entries):
                    data_uri = _process(img)
                    if not data_uri:
                        skipped += 1
                        continue
                    _register(sku, data_uri)
                # Any extra images beyond SKU count are ignored
            else:
                for _r, _c, img in img_entries:
                    data_uri = _process(img)
                    if not data_uri:
                        skipped += 1
                        continue
                    for sku in skus:
                        _register(sku, data_uri)

    msg = f'      images:     {len(images)} SKUs mapped'
    if skipped:
        msg += f' ({skipped} skipped)'
    print(msg)
    return images



# -------------------------------------------------------------
# Builders for the 4 JSON blocks
# -------------------------------------------------------------
def load_pd_config():
    """Load Monthly PD Report/pd_table_config.json. Used for ASI exclusion etc.
    Returns empty dict if file missing."""
    config_path = MONTHLY_DIR / 'pd_table_config.json'
    if not config_path.exists():
        return {}
    return json.loads(config_path.read_text(encoding='utf-8'))


def compute_mp_set(tracker_rows, config=None):
    """Return set of SKUs whose Weekly Tracker Current Status is MP or Inspection.
    Both stages count as 'Project Released' (Summer 2026-05-04: 'Inspection 和 MP
    是一样的').

    5/12: union with config['mp_overrides'] — PD Table SKUs to force-treat as MP
    when Tracker collapses multiple color variants into one MP row.
    """
    released_statuses = {'MP', 'INSPECTION'}
    base = {r['sku'] for r in tracker_rows
            if (r.get('current_status') or '').strip().upper() in released_statuses}
    overrides = set((config or {}).get('mp_overrides', []))
    return base | overrides


def brand_for_sku(sku, pd_brand=''):
    """Page 1 Brand filter 的取值（Summer 2026-08-04 定）。

    **PD Table 的 Brand 列优先**——有值就用它，只做大小写归一（源数据里
    Chefman / CHEFMAN 混用、Chef IQ / Chef iQ 混用，不归一会在筛选器里
    变成两个互相分裂的选项）。
    PD Table 该列为空时才回落到 SKU 规则：以 `CQ` 开头的、以及 `C60`，
    归 Chef iQ；其余 Chefman。（例：CQ60 V4 在 PD Table 里 Brand 是空的）
    """
    raw = (pd_brand or '').strip()
    if raw:
        u = raw.upper().replace(' ', '')
        if 'IQ' in u:
            return 'Chef iQ'
        if 'CHEFMAN' in u:
            return 'Chefman'
        return raw          # 出现没见过的品牌值时原样保留，不静默改写
    s = (sku or '').strip().upper()
    if s.startswith('CQ') or s == 'C60':
        return 'Chef iQ'
    return 'Chefman'


def build_page1_data(pd_main, tracker_rows, asi_set, mp_set, images=None):
    """Page 1 = product cards. Driven by PD Table main rows.

    Joins each PD Table SKU with Tracker (for status/risk/crd/pm) by exact match.
    SKIPS SKUs that have no PD Table entry (no commercial info → can't draw card).
    SKIPS SKUs in `asi_set` (After Sales Improvement — Page 2/3 only) and `mp_set`
    (already MP — Project Released, separate stat card).

    `images` is an optional dict {sku → base64 data URI} produced by
    extract_sku_images(); SKUs without an image fall through to the placeholder
    icon in the template.
    """
    tracker_by_sku = {row['sku']: row for row in tracker_rows}
    if images is None:
        images = {}

    # Order: PM section order from PD Table, then row order within each section.
    items = []
    for section in PM_SECTION_ORDER:
        # Collect PD Table rows for this section, preserving insertion order
        section_skus = [(sku, rec) for sku, rec in pd_main.items() if rec['pm_section'] == section]
        for sku, rec in section_skus:
            tr = tracker_by_sku.get(sku)

            # Format cost: ensure $ prefix
            cost = rec.get('cost', '')
            if cost and not cost.startswith('$'):
                # Add $ if numeric-ish
                if any(c.isdigit() for c in cost):
                    cost = '$' + cost.lstrip('$ ')

            # 1 PD Table row = 1 Page 1 card (umbrella expansion removed
            # 2026-05-07: rebuild_pdtable.py now splits multi-SKU cells in
            # PD updates directly so each variant gets its own PD Table row.
            # ASI / MP exclusion: ASI lives in config; MP is auto-detected
            # from Tracker Current Status.
            if sku in asi_set or sku in mp_set:
                continue

            # Page 1 uses PD Table as the authoritative source for tier and
            # category. If a SKU has empty tier/category in PD Table, it's
            # typically a US-side project that China is only tracking (not
            # commercializing) — those are intentionally hidden from the
            # Sales card view.
            item = {
                'sku': sku,
                'category': normalize_category(rec.get('category', '')),
                'tier': rec.get('tier', ''),
                'brand': brand_for_sku(sku, rec.get('brand', '')),
                'description': rec.get('description', ''),
                'topFeature': rec.get('top_feature', ''),
                'uf1': rec.get('uf1', ''),
                'uf2': rec.get('uf2', ''),
                'uf3': rec.get('uf3', ''),
                'msrp': rec.get('msrp', ''),
                'sampleETA': rec.get('sample_eta', ''),
                'poPlaced': rec.get('po_placed', ''),
                # PO status for Page 1 filter/badge/detail now reads the
                # Weekly Tracker PO/订单状态 col (col N) via parse_po — same
                # source as Page 3. PD Table 'PO Placed?' (po_placed) is
                # mostly empty/stale, so the homepage PO filter must use
                # the Tracker. (2026-06-01 fix)
                # Page 1 homepage is 2-state: INTENT collapses into No PO
                # (only confirmed PO = PO Placed). Page 3 keeps INTENT distinct.
                'poStatus': ('NO' if parse_po(tr['po_status'])[0] == 'INTENT'
                             else parse_po(tr['po_status'])[0]) if tr else '',
                'poBuyer': parse_po(tr['po_status'])[1] if tr else '',
                'estInspection': rec.get('est_inspection', ''),
                'factory': rec.get('factory', ''),
                'market': rec.get('market', ''),
                'cost': cost,
                'buffer': rec.get('buffer', ''),
                'port': rec.get('port', ''),
                'duty': rec.get('duty', ''),
                'hc40': rec.get('hc40', ''),
                'compModel': rec.get('comp_model', ''),
                'rjDiff': rec.get('rj_diff', ''),
                'note1': rec.get('note1', ''),
                'note2': rec.get('note2', ''),
                'pmSection': section,
                'currentStatus': tr['current_status'] if tr else '',
                'risk': tr['risk'] if tr else '',
                'crd': tr['crd'] if tr else '',
                'pm': tr['pm'] if tr else '',
                'pv': tr.get('pv', '') if tr else '',
                # Embedded base64 thumbnail (empty falls back to placeholder
                # icon in template). _sku_image_aliases() handles parenthetical
                # color codes — e.g. images['RJ50-SFDAF-25D'] aliases to
                # 'RJ50-SFDAF-25D(SS)' so SS/BLK siblings can share one image.
                'image': images.get(sku, ''),
            }
            items.append(item)
    return items



def build_placeholder_cards(tracker_rows, pd_main, asi_set, mp_set, images=None):
    """Cards for SKUs that exist in Tracker but not in PD Table — PM hasn't
    filled commercial info yet. Renders as a dashed "Pending PM Input" card on
    Page 1 so PMO/Sales can still see they exist + count them in stats. ASI
    and MP SKUs are skipped (ASI = not on Page 1 by design; MP = already
    Released).
    """
    images = images or {}
    pd_skus = set(pd_main.keys())
    items = []
    for row in tracker_rows:
        sku = row.get('sku') or ''
        if not sku:
            continue
        if sku in pd_skus or sku in asi_set or sku in mp_set:
            continue
        items.append({
            'sku': sku,
            'category': normalize_category(row.get('category') or ''),
            'tier': row.get('tier') or '',
            'pm': row.get('pm') or '',
            'pmSection': row.get('pm_section') or '',
            'currentStatus': row.get('current_status') or '',
            'risk': row.get('risk') or '',
            'crd': row.get('crd') or '',
            'issue': row.get('issue') or '',
            'nextAction': row.get('next_action') or '',
            'image': images.get(sku, ''),
            'isPlaceholder': True,
            # Empty commercial fields (modal renders banner instead)
            'description': '', 'topFeature': '', 'uf1': '', 'uf2': '', 'uf3': '',
            'msrp': '', 'sampleETA': '', 'poPlaced': '',
            'estInspection': '', 'factory': '', 'market': '',
            'cost': '', 'buffer': '', 'port': '', 'duty': '', 'hc40': '',
            'compModel': '', 'rjDiff': '', 'note1': '', 'note2': '',
            'brand': brand_for_sku(sku),
            'onProjectList': False,
        })
    return items


def build_page3_data(tracker_rows, asi_set=None, pd_main_skus=None, mp_set=None):
    """Page 3 = Weekly Tracker rows. One per Tracker SKU.

    asi_set: SKUs flagged as After Sales Improvement → tag `isASI=true` so the
        NPD/ASI filter can hide them.
    pd_main_skus: set of SKUs that exist in PD Table (umbrella form). Used to
        tag `inPdTable=true` so the Risk Detail Panel can mirror the stat-card
        filter (stat counts only PD-Table-backed projects).
    mp_set: SKUs whose Tracker status is MP. Tagged `isMP=true` so the Risk
        Detail Panel can exclude already-released projects.
    """
    asi_set = asi_set or set()
    pd_main_skus = pd_main_skus or set()
    mp_set = mp_set or set()
    items = []
    for i, row in enumerate(tracker_rows, 1):
        po_status, po_buyer = parse_po(row['po_status'])
        items.append({
            'num': str(i),
            'sku': row['sku'],
            'pv': row.get('pv', ''),   # 'Parent' / 'Variant' / '' (Tracker C col, added 5/6)
            'category': row['category'],
            'risk': row['risk'],
            'pm': row['pm'],
            'tier': row['tier'],
            'lastUpdate': row['last_update'],
            'currentStatus': row['current_status'],
            'issue': row['issue'],
            'nextAction': row['next_action'],
            'crd': row['crd'],
            'location': infer_location(row['next_action']),
            'pmSection': row['pm_section'],
            'stages': row['stages'],
            'paStatus': row.get('pa_status', ''),  # 5/12: PA (Product Authorization) signed status
            'sixA': row.get('six_a', ''),  # 6/15: 6A (Amazon SIOC packaging cert) status
            'poStatus': po_status,
            'poBuyer': po_buyer,
            'poBuyers': parse_po_buyers(row['po_status'])[1],  # per-buyer rows for Page 3 PO cell
            'poRaw': row['po_status'],  # preserved for hover/tooltip if needed
            'isASI': row['sku'] in asi_set,
            'inPdTable': row['sku'] in pd_main_skus,
            'isMP': row['sku'] in mp_set,
        })
    return items


def build_pipeline_data(tracker_rows, asi_set=None):
    """Page 2 = Pipeline Timeline. 12 stages, projects grouped by current stage.

    asi_set: optional set of SKUs flagged as After Sales Improvement; each
    project is tagged `isASI=true` so the front-end NPD/ASI filter can hide
    them and the on-page count badges can be recomputed on filter change.
    """
    asi_set = asi_set or set()
    counts = [0] * len(PIPELINE_LABELS)
    projects = [[] for _ in PIPELINE_LABELS]

    for row in tracker_rows:
        status = row['current_status']
        idx = STATUS_TO_PIPELINE.get(status)
        # Try case-insensitive / contains match for fallback
        if idx is None and status:
            sl = status.lower()
            for k, v in STATUS_TO_PIPELINE.items():
                if k.lower() == sl:
                    idx = v
                    break
        if idx is None:
            continue  # status doesn't map to a pipeline stage (e.g., '色样确认中', '—')
        counts[idx] += 1
        po_status, po_buyer = parse_po(row['po_status'])
        projects[idx].append({
            'sku': row['sku'],
            'category': row['category'],
            'pm': row['pm'],
            'risk': row['risk'],
            'status': PIPELINE_LABELS[idx],
            'action': row['next_action'],
            'poStatus': po_status,
            'poBuyer': po_buyer,
            'paStatus': row.get('pa_status', ''),
            'sixA': row.get('six_a', ''),
            'isASI': row['sku'] in asi_set,
        })

    return {'counts': counts, 'labels': PIPELINE_LABELS, 'projects': projects}


def build_summary_stats(page1, tracker_rows, asi_set, mp_set, pd_main_skus):
    """Auto-compute the 5 stats bar numbers.

    Rules:
    - Total Projects: NPD + ASI active dev (only MP excluded). page1 is
      filtered to exclude both ASI and MP for cards, so add ASI-non-MP back.
    - High Risk / Medium Risk: Tracker rows where risk matches AND the SKU is
      in PD Table AND it's not ASI AND it's not MP. (Same filter as the Risk
      Detail Panel — stat number == panel row count by construction.) Counted
      per Tracker row (umbrella) so an umbrella with 4 variants is 1, not 4.
    - Tier 1 (CSM): all T1 in Tracker including MP T1 (Summer's exception).
    - Project Released: total MP count.
    """
    visible = [p for p in page1 if p.get('category')]
    # 5/12: stats.total = visible page1 cards exactly (Summer feedback —
    # "stats == 可见卡片"; ASI active dev shown only on Page 2/3, not counted)
    total = len(visible)

    def _is_panel_visible(r):
        # Now inclusive of "Tracker has it but PD Table doesn't" — those
        # render as placeholder cards on Page 1, so they should also count
        # toward the High/Mid stats and the Risk Detail Panel.
        sku = r.get('sku')
        if not sku: return False
        if sku in asi_set: return False
        if sku in mp_set: return False
        return True

    high = sum(1 for r in tracker_rows if _is_panel_visible(r) and r.get('risk') == '高')
    mid = sum(1 for r in tracker_rows if _is_panel_visible(r) and r.get('risk') == '中')
    t1 = sum(1 for r in tracker_rows if (r.get('tier') or '').strip() == '1')
    released = len(mp_set)
    return {'total': total, 'high': high, 'mid': mid, 't1': t1, 'released': released}


def build_released_data(tracker_rows, mp_set, pd_main=None):
    """Data for the 'Project Released' stat card dropdown.
    Columns: SKU / PM / Category / PO info / CRD.

    5/12: iterate mp_set so dropdown count == stats.released. Dedupe by SKU
    (alias collisions like 9TW-V3→V2 used to produce duplicate rows). For
    mp_overrides SKUs without a Tracker row, synthesize an entry using PD
    Table category if available.
    """
    pd_main = pd_main or {}
    tracker_by_sku = {}
    for r in tracker_rows:
        sku = r['sku']
        existing = tracker_by_sku.get(sku)
        # Prefer MP-status row over non-MP for duplicates from alias collision
        if existing is None or (r.get('current_status', '').strip().upper() == 'MP'
                                and existing.get('current_status', '').strip().upper() != 'MP'):
            tracker_by_sku[sku] = r
    items = []
    for sku in sorted(mp_set):
        r = tracker_by_sku.get(sku)
        if r:
            po_status, po_buyer = parse_po(r.get('po_status', ''))
            items.append({
                'sku': sku,
                'pm': r.get('pm', ''),
                'category': r.get('category', ''),
                'poStatus': po_status,
                'poBuyer': po_buyer,
                'poRaw': r.get('po_status', ''),
                'crd': r.get('crd', ''),
            })
        else:
            # mp_overrides SKU — no Tracker row; pull category from PD Table
            rec = pd_main.get(sku, {})
            items.append({
                'sku': sku,
                'pm': rec.get('pm_section', '').split(' — ')[0] if rec.get('pm_section') else '',
                'category': rec.get('category', ''),
                'poStatus': '',
                'poBuyer': '',
                'poRaw': '',
                'crd': '',
            })
    return items


# -------------------------------------------------------------
# Location heuristic — infer from Next Action keywords
# -------------------------------------------------------------
# US-side activities: Culinary review, design/artwork iterations, US-team confirmations
US_KEYWORDS = [
    'Culinary', 'culinary',
    'design', 'Design',
    'artwork', 'Artwork',
    'confirm', 'Confirm', '确认',  # confirm often means waiting on US sign-off
    'Andrew', 'Ryan',              # US team contacts
    'packaging', 'Packaging',
    'Sales', 'sales',
    'Pantone',
]
# China-side activities: prototyping, factory milestones, performance/life testing
CHINA_KEYWORDS = [
    '手板', '打样', '样品',
    'EB', 'PP', 'FOT',
    '寿命', '性能', '测试',
    '装配', '模具', '工厂',
    '大货', '色样', '注塑', '钣金',
    '量产', '试产',
]


def infer_location(next_action):
    """Infer China/US/Both/'' from Next Action text using keyword heuristics."""
    if not next_action:
        return ''
    has_us = any(kw in next_action for kw in US_KEYWORDS)
    has_cn = any(kw in next_action for kw in CHINA_KEYWORDS)
    if has_us and has_cn:
        return 'Both'
    if has_us:
        return 'US'
    if has_cn:
        return 'China'
    return ''


# -------------------------------------------------------------
# PO parsing — derive (status, buyer) from Tracker C11 free-text
# -------------------------------------------------------------
# Phrases that mean "no PO yet" / cancelled
PO_NEGATIVE_PHRASES = [
    '暂无订单', '无Open PO', '无open PO', '无PO', '无 PO',
    '项目Pending,无Open PO', '项目取消',
]
# 2026-06-01 (Summer rule): intent / inquiry / not-yet-placed are NOT a placed
# PO. They get a distinct 'INTENT' status so Page 3 shows three buckets
# (无PO / XX意向 / XX已PO). Page 1 homepage collapses INTENT into No PO.
PO_INTENT_PHRASES = ['意向', '询单', '即将']
# Known buyer keywords. ORDER MATTERS — longer multi-word names listed first
# so "Canadian Tire" matches before "Canadian", "Walmart 3P" can be normalized
# to "Walmart" by listing "Walmart" alone (after multi-word variants).
PO_BUYER_KEYWORDS = [
    'Canadian Tire',
    "Sam's", 'Sams Club', 'Sams',
    "Kohl's",
    'PriceSmart',
    'Loblaws',
    'Menards',
    'Costco',
    'Walmart',
    'Amazon',
    'Target',
    'AAFES', 'Macy', 'BJ', 'AMZ',
    # Markets / channels (keep last so they don't shadow customer names)
    'MX', 'CA', 'EU', 'UK',
]


# -------------------------------------------------------------
# Translation helpers (Chinese → English for US Sales version)
# -------------------------------------------------------------
def load_translations():
    """Load Chinese→English dictionary. Returns empty dict if file missing."""
    if not TRANSLATIONS_PATH.exists():
        return {}
    import json as _json
    with open(TRANSLATIONS_PATH, 'r', encoding='utf-8') as f:
        return _json.load(f)


def translate(text, trans_dict):
    """Look up translation. If not found, return original (and we'll log it)."""
    if not text:
        return text
    return trans_dict.get(text, text)


# Risk display strings for English version (中文 risk values mapped to badges in JS,
# but for risk filter dropdown values and any raw display we map them here too).
RISK_ZH_TO_EN = {'高': 'High', '中': 'Medium', '低': 'Low', '—': '—'}


def translate_page1(items, trans_dict):
    """Translate page1 cards. Fields with potential Chinese: currentStatus, category, crd, issue, nextAction (placeholder cards carry tracker issue/nextAction, shown in card detail + risk table)."""
    out = []
    for p in items:
        new_p = dict(p)
        for key in ['currentStatus', 'category', 'crd', 'issue', 'nextAction']:
            if p.get(key):
                new_p[key] = translate(p[key], trans_dict)
        out.append(new_p)
    return out


def translate_page3(items, trans_dict):
    """Translate page3 (Weekly Tracker) rows: issue / nextAction / currentStatus / category / poRaw / crd."""
    out = []
    for p in items:
        new_p = dict(p)
        for key in ['issue', 'nextAction', 'currentStatus', 'category', 'poRaw', 'crd', 'paStatus', 'sixA']:
            if p.get(key):
                new_p[key] = translate(p[key], trans_dict)
        out.append(new_p)
    return out


def translate_pipeline(pipe, trans_dict):
    """Translate pipeline projects' action and category fields."""
    new_pipe = {
        'counts': pipe['counts'],
        'labels': pipe['labels'],
        'projects': [],
    }
    for stage_projs in pipe['projects']:
        new_stage = []
        for proj in stage_projs:
            np = dict(proj)
            if proj.get('action'):
                np['action'] = translate(proj['action'], trans_dict)
            if proj.get('category'):
                np['category'] = translate(proj['category'], trans_dict)
            if proj.get('paStatus'):
                np['paStatus'] = translate(proj['paStatus'], trans_dict)
            if proj.get('sixA'):
                np['sixA'] = translate(proj['sixA'], trans_dict)
            new_stage.append(np)
        new_pipe['projects'].append(new_stage)
    return new_pipe


def translate_released(items, trans_dict):
    """Translate released-panel rows: category / poRaw / crd.
    (2026-07-14 加：EN 版此前把 released 整包原样注入，中文 CRD/PO 原文漏进英文版。)"""
    out = []
    for p in items:
        new_p = dict(p)
        for key in ['category', 'poRaw', 'crd']:
            if p.get(key):
                new_p[key] = translate(p[key], trans_dict)
        out.append(new_p)
    return out


def report_untranslated_flat(items, keys, trans_dict):
    """Generic missing-translation check over a flat list of dicts（2026-07-14 加，
    补 released / crd_changes 两个此前不在告警覆盖内的数据块）。"""
    import re as _re
    zh_pat = _re.compile(r'[一-鿿]')
    out = set()
    for p in items:
        for k in keys:
            v = p.get(k)
            if v and zh_pat.search(str(v)) and str(v) not in trans_dict:
                out.add(str(v))
    return out


def report_untranslated(items_p1, items_p3, items_pipe, trans_dict):
    """Walk all data and report which Chinese strings have no translation."""
    import re as _re
    zh_pat = _re.compile(r'[一-鿿]')
    untrans = set()
    # page1
    for p in items_p1:
        for k in ['currentStatus', 'category', 'description', 'topFeature', 'crd', 'issue', 'nextAction']:
            v = p.get(k, '')
            if v and zh_pat.search(str(v)) and v not in trans_dict:
                untrans.add(v)
    # page3
    for p in items_p3:
        for k in ['issue', 'nextAction', 'currentStatus', 'category', 'poRaw', 'crd', 'paStatus', 'sixA']:
            v = p.get(k, '')
            if v and zh_pat.search(str(v)) and v not in trans_dict:
                untrans.add(v)
    # pipeline
    for stage in items_pipe['projects']:
        for p in stage:
            for kk in ('action', 'category', 'paStatus', 'sixA'):
                v = p.get(kk, '')
                if v and zh_pat.search(str(v)) and v not in trans_dict:
                    untrans.add(v)
    return untrans


def parse_po(po_status_text):
    """Parse Tracker PO/订单状态 text → (status, buyer).

    Returns:
      status: 'YES' | 'INTENT' | 'NO' | ''  (placed / intent-only / no PO / unknown)
      buyer:  string or ''        (extracted customer/channel name)
    """
    if not po_status_text:
        return '', ''
    text = po_status_text.strip()
    if not text:
        return '', ''
    # Negative phrases → No PO
    if any(neg in text for neg in PO_NEGATIVE_PHRASES):
        return 'NO', ''
    # Extract buyer / channel name (used by both INTENT and YES)
    buyer = ''
    for kw in PO_BUYER_KEYWORDS:
        if kw in text:
            buyer = kw
            break
    if not buyer:
        m = re.search(r'\bfor\s+([A-Z][A-Za-z\']+)', text)
        if m:
            buyer = m.group(1)
    # Intent / inquiry / about-to-place → INTENT (not a placed PO). Surfaced as
    # its own bucket on Page 3 (XX意向); Page 1 collapses it into No PO.
    if any(ph in text for ph in PO_INTENT_PHRASES):
        return 'INTENT', buyer
    # Confirmed PO: buyer keyword, "for X", or other substantive text
    # (e.g. '已下PO', bare channel name) → placed.
    return 'YES', buyer


def parse_po_buyers(po_status_text):
    """Parse Tracker PO text into per-buyer rows for the Page 3 PO cell.

    Splits the free-text on ; ； , separators and, for each segment, derives a
    status (YES/INTENT) and the buyer name (segment text minus PO-status tokens).
    Captures buyers not in PO_BUYER_KEYWORDS (e.g. Meijer, Amazon CA).

    Returns (overall_status, buyers_list) where buyers_list is a list of
    {'buyer': str, 'status': 'YES'|'INTENT'}. Returns (overall, []) when the
    overall status is NO / unknown, so single-buyer and No-PO rows render exactly
    as before via the poBadge fallback.
    """
    overall = parse_po(po_status_text)[0]
    if overall not in ('YES', 'INTENT'):
        return overall, []
    strip_tokens = ['已下单', '已下PO', '已PO', '下PO', '有Open PO', '有open PO',
                    '有PO', '意向', '询单', '即将', '订单', 'PO', '已下']
    buyers = []
    seen = set()
    for seg in re.split(r'[;；,，]', po_status_text):
        seg = seg.strip()
        if not seg:
            continue
        if any(neg in seg for neg in PO_NEGATIVE_PHRASES):
            continue  # skip pure "no PO" fragments
        st = 'INTENT' if any(ph in seg for ph in PO_INTENT_PHRASES) else 'YES'
        buyer = seg
        for tok in strip_tokens:
            buyer = buyer.replace(tok, '')
        buyer = buyer.strip(' :：-—_/、()（）')
        key = (buyer, st)
        if key in seen:
            continue
        seen.add(key)
        buyers.append({'buyer': buyer, 'status': st})
    return overall, buyers


# Map each PM section header → the English category names Sales would recognize.
# Banner shows these names instead of the messy raw category strings from PD Table.
PM_SECTION_TO_CATEGORIES = {
    'Cottee Wei — 空气炸锅 + T1 项目':
        ['Air Fryers'],
    'Rowling Luo — 烤箱 / 面包机 / 饭煲 / 慢炖锅 / 油炸锅':
        ['Ovens', 'Bread Maker', 'Rice Cooker', 'Slow Cooker', 'Deep Fryer'],
    'Serena Sun — ICEMAN / 咖啡 / 冰淇淋':
        ['Iceman', 'Coffee', 'Ice Cream'],
    'Chris Zhou — 烤盘 / 搅拌类 + MX 项目':
        ['Griddle', 'Blender'],
    'Liz Liu — 水壶 + 微波炉':
        ['Kettle', 'Microwave'],
    'Jenifer Yuan — CSM 杭州 (C60 / C45 / CQ60 / C22)':
        ['Thermometer', 'Vacuum Sealer'],
}

# How many pending SKUs a PM must have before banner flags their categories.
# Singletons are usually one-off SKU-level issues, not a category-wide data gap.
BANNER_PM_THRESHOLD = 3


def build_banner_html(tracker_rows, pd_main, asi_set, mp_set):
    """Detect PMs with systemic data gaps and surface their domain categories.

    Logic (Phase 2 — adapted to new pure-mirror SOP):
      - For each Tracker SKU that is NOT in PD Table, NOT in ASI list, NOT MP:
        this is a gap — PM owes commercial info in next PD updates.
      - Count gaps per PM section. If ≥ BANNER_PM_THRESHOLD, flag categories.
      - Singletons (1-2 missing SKUs) are treated as SKU-level specifics and
        not surfaced — they're typically intentional gaps.

    Returns: HTML string for banner block (empty if no PM hits threshold).
    """
    pd_skus = set(pd_main.keys())
    pm_pending_count = {}
    for r in tracker_rows:
        sku = r.get('sku')
        if not sku:
            continue
        if sku in pd_skus or sku in asi_set or sku in mp_set:
            continue
        section = r.get('pm_section') or ''
        if section:
            pm_pending_count[section] = pm_pending_count.get(section, 0) + 1

    # Collect categories to flag
    flagged_cats = []
    flagged_pms = []
    for section in PM_SECTION_ORDER:
        if pm_pending_count.get(section, 0) >= BANNER_PM_THRESHOLD:
            cats = PM_SECTION_TO_CATEGORIES.get(section, [])
            flagged_cats.extend(cats)
            # Extract PM short name (first 2 words before em dash)
            pm_short = section.split('—')[0].strip()
            flagged_pms.append(pm_short)

    if not flagged_cats:
        return ''

    # Format category list: "A, B and C" English style
    if len(flagged_cats) == 1:
        cat_str = flagged_cats[0]
    elif len(flagged_cats) == 2:
        cat_str = f'{flagged_cats[0]} and {flagged_cats[1]}'
    else:
        cat_str = ', '.join(flagged_cats[:-1]) + f' and {flagged_cats[-1]}'

    banner = (
        f'<div class="data-banner">'
        f'<span class="banner-icon">⚠️</span>'
        f'<div><strong>Data note:</strong> '
        f'{cat_str} categor{"y" if len(flagged_cats) == 1 else "ies"} '
        f'currently lack complete commercial data — pending updates from related PM.</div>'
        f'</div>'
    )
    return banner


# -------------------------------------------------------------
# Render & rotate
# -------------------------------------------------------------
def render_template(template_text, page1, pipeline_us, pipeline_mx, page3, stats, banner, released, po_labels=None, pa_labels=None, sixa_labels=None, data_asof='', crd_changes=None, pa6a_labels=None, ui_labels=None):
    """Substitute placeholders with JSON / HTML."""
    # TEST 2026-08-10: Region toggle 等新增 UI 文案(中英各一套)
    if ui_labels is None:
        ui_labels = {
            'region_cn': 'China', 'region_us': 'US',
            'all_categories': 'All Categories', 'all_pms': 'All PMs',
            'all_tiers': 'All Tiers',
            'us_status': 'Status (US PD)',
            'stats_note': 'Stats above cover China PD projects only — '
                          'switch the Region toggle below to view US PD projects.',
        }
    if po_labels is None:
        po_labels = {'placed': 'PO Placed', 'intent': 'Intent', 'noPO': 'No PO'}
    if pa_labels is None:
        pa_labels = {'no': 'No', 'ongoing': 'Ongoing', 'waiting': 'Waiting for Signing', 'done': 'Done'}
    if sixa_labels is None:
        sixa_labels = {'no': 'No', 'ongoing': 'Ongoing', 'done': 'Done', 'na': 'Not Needed'}
    if pa6a_labels is None:
        pa6a_labels = {'tile': 'PA / 6A Incomplete', 'panel': 'PA / 6A Incomplete · PP/MP+'}
    out = template_text
    out = out.replace('{{DATA_ASOF}}', data_asof)
    out = out.replace('{{PO_LABELS}}', json.dumps(po_labels, ensure_ascii=False))
    out = out.replace('{{PA_LABELS}}', json.dumps(pa_labels, ensure_ascii=False))
    out = out.replace('{{SIXA_LABELS}}', json.dumps(sixa_labels, ensure_ascii=False))
    out = out.replace('{{PA6A_LABELS}}', json.dumps(pa6a_labels, ensure_ascii=False))
    out = out.replace('{{PAGE1_DATA}}', json.dumps(page1, ensure_ascii=False))
    out = out.replace('{{PIPELINE_US_DATA}}', json.dumps(pipeline_us, ensure_ascii=False))
    out = out.replace('{{PIPELINE_MX_DATA}}', json.dumps(pipeline_mx, ensure_ascii=False))
    out = out.replace('{{PAGE3_DATA}}', json.dumps(page3, ensure_ascii=False))
    out = out.replace('{{SUMMARY_STATS}}', json.dumps(stats, ensure_ascii=False))
    out = out.replace('{{RELEASED_DATA}}', json.dumps(released, ensure_ascii=False))
    out = out.replace('{{CRD_CHANGE_DATA}}', json.dumps(crd_changes or [], ensure_ascii=False))
    out = out.replace('{{BANNER_BLOCK}}', banner)
    out = out.replace('{{REPORT_PERIOD}}', REPORT_PERIOD)
    out = out.replace('{{REGION_CN_LABEL}}', ui_labels['region_cn'])
    out = out.replace('{{REGION_US_LABEL}}', ui_labels['region_us'])
    out = out.replace('{{ALL_CATEGORIES_LABEL}}', ui_labels['all_categories'])
    out = out.replace('{{ALL_PMS_LABEL}}', ui_labels['all_pms'])
    out = out.replace('{{ALL_TIERS_LABEL}}', ui_labels['all_tiers'])
    out = out.replace('{{US_STATUS_LABEL}}', ui_labels['us_status'])
    out = out.replace('{{STATS_NOTE}}', json.dumps(ui_labels['stats_note'], ensure_ascii=False))
    # Sanity: no placeholders should remain
    leftover = re.findall(r'\{\{[A-Z_]+\}\}', out)
    if leftover:
        raise RuntimeError(f'Unfilled placeholders: {leftover}')
    return out


def write_with_rotation(html_text, out_path, prev_path):
    """Apply rotation rule, then write new file safely."""
    if out_path.exists():
        try:
            shutil.move(str(out_path), str(prev_path))
            print(f'  rotation: {out_path.name} -> {prev_path.name}')
        except PermissionError:
            shutil.copyfile(str(out_path), str(prev_path))
            print(f'  rotation: copied {out_path.name} -> {prev_path.name}')

    scratch_file = SCRATCH / out_path.name
    scratch_file.write_text(html_text, encoding='utf-8')
    try:
        shutil.copyfile(str(scratch_file), str(out_path))
    except PermissionError:
        out_path.write_text(html_text, encoding='utf-8')
    print(f'  wrote: {out_path.name} ({len(html_text):,} chars)')


def main():
    print('=== China PD Monthly Report Builder ===')
    print(f'Output (CN): {OUT_NAME}')
    print(f'Output (EN): {OUT_NAME_EN}')
    print()

    print(f'[1/5] Loading template: {TEMPLATE_PATH.name}')
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f'Template missing: {TEMPLATE_PATH}')
    template = TEMPLATE_PATH.read_text(encoding='utf-8')

    print(f'[2/5] Loading data sources')
    # Load config early so load_tracker can apply sku_aliases (5/12)
    config = load_pd_config()
    sku_aliases = config.get('sku_aliases', {})
    if sku_aliases:
        print(f'      sku_aliases applied: {len(sku_aliases)} mapping(s)')
    print(f'      tracker:    {TRACKER_PATH.name}')
    tracker_rows = load_tracker(TRACKER_PATH, sku_aliases=sku_aliases)
    print(f'                  -> {len(tracker_rows)} SKU rows')

    print(f'      pd table:   {PDTABLE_PATH.name}')
    pd_main, pd_pending = load_pd_table(PDTABLE_PATH)
    print(f'                  -> {len(pd_main)} main SKUs, {len(pd_pending)} pending')

    if PDUPDATES_PATH:
        print(f'      pd updates: {PDUPDATES_PATH.name} (extracting product images)')
    else:
        print(f'      pd updates: NONE FOUND (no embedded images this run)')
    images = extract_sku_images(PDUPDATES_PATH)

    print(f'[3/5] Building data blocks')
    # config already loaded above (needed for sku_aliases in load_tracker)
    # 5/19: ASI source of truth is Tracker col E (npd_asi field), not config.
    # Blank col E = NPD by default; only 'ASI' marker excludes from Page 1.
    asi_set = {r['sku'] for r in tracker_rows if (r.get('npd_asi') or '').strip().upper() == 'ASI'}
    mp_set = compute_mp_set(tracker_rows, config)
    print(f'      ASI exclusion: {len(asi_set)} SKUs from Tracker col E: {sorted(asi_set)}')
    print(f'      MP/Released set: {len(mp_set)} SKUs (Tracker MP + config mp_overrides)')
    page1 = build_page1_data(pd_main, tracker_rows, asi_set, mp_set, images)
    placeholders = build_placeholder_cards(tracker_rows, pd_main, asi_set, mp_set, images)
    page1.extend(placeholders)

    cn_card_count = len(page1)

    # TEST (2026-08-10): US PM projects (both sheets) join Page 1 as a second
    # region. The homepage card area gets a China / US toggle (default China);
    # the top stats bar stays China-only — US projects have no Weekly Tracker
    # row, so risk / CRD / PA·6A / PO have no value for them.
    us_items, _ = load_us_pd()
    report_us_overlap(us_items, pd_main, tracker_rows)
    us_items = drop_cn_pm_projects(us_items, tracker_rows, pd_main)

    # Stats computed BEFORE the US append so the top bar stays China-only.
    stats = build_summary_stats(page1, tracker_rows, asi_set, mp_set, set(pd_main.keys()))

    page1.extend(us_items)

    # Merge small categories (< SMALL_CAT_THRESHOLD cards) into "Other".
    # Counted PER REGION (2026-08-10): the two regions render as separate views,
    # so a China category with 1 card must still collapse even if the US side
    # has 4 of the same kind (e.g. Thermometer: CN 1 + US 4).
    SMALL_CAT_THRESHOLD = 3
    from collections import Counter as _Counter
    moved = 0
    small_total = 0
    for region in ('CN', 'US'):
        bucket = [p for p in page1 if (p.get('source') == 'US') == (region == 'US')]
        cat_counts = _Counter(p.get('category', '') for p in bucket)
        small_cats = {c for c, n in cat_counts.items() if n < SMALL_CAT_THRESHOLD}
        small_total += len(small_cats)
        for p in bucket:
            if p.get('category', '') in small_cats:
                p['category'] = 'Other'
                moved += 1
    print(f'      page1Data: {len(page1)} cards '
          f'(CN {cn_card_count} = {cn_card_count-len(placeholders)} PD Table + {len(placeholders)} placeholders'
          f' | US {len(us_items)}); {moved} moved to "Other" from {small_total} small cats')
    page3 = build_page3_data(tracker_rows, asi_set, set(pd_main.keys()), mp_set)
    print(f'      page3Data: {len(page3)} tracker rows')
    us_rows = [r for r in tracker_rows if not is_mx_sku(r.get('sku', ''))]
    mx_rows = [r for r in tracker_rows if is_mx_sku(r.get('sku', ''))]
    pipeline_us = build_pipeline_data(us_rows, asi_set)
    pipeline_mx = build_pipeline_data(mx_rows, asi_set)
    print(f'      pipelineUSData: counts={pipeline_us["counts"]} (total={sum(pipeline_us["counts"])})')
    print(f'      pipelineMXData: counts={pipeline_mx["counts"]} (total={sum(pipeline_mx["counts"])})')
    crd_changes = compute_crd_changes(tracker_rows, config, sku_aliases)
    stats['crd'] = len(crd_changes)
    print(f'      summaryStats: {stats}')
    released = build_released_data(tracker_rows, mp_set, pd_main)
    print(f'      releasedData: {len(released)} entries (Project Released dropdown)')

    banner = build_banner_html(tracker_rows, pd_main, asi_set, mp_set)
    if banner:
        print(f'[4/5] Banner ON')
    else:
        print(f'[4/5] Banner OFF')

    print(f'[5/5] Render + rotate (Chinese)')
    html_out = render_template(template, page1, pipeline_us, pipeline_mx, page3, stats, banner, released, po_labels={'placed': '已PO', 'intent': '意向', 'noPO': '无PO'}, pa_labels={'no': '未申请', 'ongoing': '申请中', 'waiting': '待批复', 'done': '已完成'}, sixa_labels={'no': '未申请', 'ongoing': '申请中', 'done': '已完成', 'na': '无需'}, data_asof=DATA_ASOF_CN, crd_changes=crd_changes, pa6a_labels={'tile': 'PA / 6A 未完成', 'panel': 'PA / 6A 未完成 · 已PP/MP'}, ui_labels={'region_cn': '中国', 'region_us': '美国', 'all_categories': '全部品类', 'all_pms': '全部 PM', 'all_tiers': '全部 Tier', 'us_status': '美方状态 (US PD)', 'stats_note': '上方统计仅含中国 PD 项目；美方项目请用下方卡片区的中国 / 美国切换查看。'})
    write_with_rotation(html_out, OUT_PATH, PREV_PATH)

    print(f'[5/5] Render + rotate (English)')
    trans = load_translations()
    print(f'      translations loaded: {len(trans)} entries')
    # CRD Change EN: override 有 reasonEN 用之，否则走翻译字典（reason 来自卡点列，与 page3 issue 同源）
    crd_changes_en = [dict(e, reason=(e.get('reasonEN') or translate(e.get('reason', ''), trans)),
                           oldCRD=translate(e.get('oldCRD', ''), trans),
                           newCRD=translate(e.get('newCRD', ''), trans)) for e in crd_changes]
    page1_en = translate_page1(page1, trans)
    page3_en = translate_page3(page3, trans)
    pipeline_us_en = translate_pipeline(pipeline_us, trans)
    pipeline_mx_en = translate_pipeline(pipeline_mx, trans)
    released_en = translate_released(released, trans)

    untranslated = report_untranslated(page1_en, page3_en, pipeline_us_en, trans)
    untranslated |= report_untranslated(page1_en, page3_en, pipeline_mx_en, trans)
    untranslated |= report_untranslated_flat(released_en, ['category', 'poRaw', 'crd'], trans)
    untranslated |= report_untranslated_flat(crd_changes_en, ['oldCRD', 'newCRD', 'reason'], trans)
    if untranslated:
        print(f'      WARNING: {len(untranslated)} Chinese strings missing translation:')
        for s in sorted(untranslated)[:10]:
            print(f'        {s[:80]!r}')
        if len(untranslated) > 10:
            print(f'        ... +{len(untranslated)-10} more')
    else:
        print(f'      OK all Chinese strings translated')

    html_out_en = render_template(template, page1_en, pipeline_us_en, pipeline_mx_en, page3_en, stats, banner, released_en, po_labels={'placed': 'PO Placed', 'intent': 'Intent', 'noPO': 'No PO'}, pa_labels={'no': 'No', 'ongoing': 'Ongoing', 'waiting': 'Waiting for Signing', 'done': 'Done'}, sixa_labels={'no': 'No', 'ongoing': 'Ongoing', 'done': 'Done', 'na': 'Not Needed'}, data_asof=DATA_ASOF_EN, crd_changes=crd_changes_en, pa6a_labels={'tile': 'PA / 6A Incomplete', 'panel': 'PA / 6A Incomplete · PP/MP+'})
    write_with_rotation(html_out_en, OUT_PATH_EN, PREV_PATH_EN)

    # GitHub Pages 首页跟随最新 EN 版（2026-07-02 加；Push_NOW.bat 的同名逻辑保留作双保险）
    index_path = MONTHLY_DIR / 'index.html'
    index_path.write_text(html_out_en, encoding='utf-8')
    print(f'  index.html synced to latest EN report')

    print('Done.')


# -------------------------------------------------------------
# TEST (2026-08-10): US PD integration — 美方要求把 US PM 的项目加进月报
# Page 1(原首页卡片页混排,不单独成页)。
#
# Source: 'US PD Update*.xlsx'(美方文件副本)。两个可见 sheet 都要收:
#   'Chef IQ' —  7 个项目(C..I 列),Chef iQ 品牌线
#   'Chefman' — 30 个项目(B..AE 列),Chefman 品牌线
# 两个 sheet 都是竖排 PD Table(行=字段,列=项目),但**行号布局不同**,且
# Chefman 多 Status 行、少 MSRP/Duty/Brand/RJ-Diff 行 → 各自一套行映射。
#
# 图: 两个 sheet 的图都是 Excel 365 "单元格内图片"(image-in-cell 富数据),
# openpyxl 的 data_only 读出来是 '#VALUE!' —— 不是没有图。用现成的
# _extract_image_in_cell_raw() 解出来:Chefman 29 张 + Chef IQ 5 张。
#
# Tier: Chefman 的 Tier 列直接读(值 1 / 1.5 / 2)。Chef IQ 的 Tier 列写的是
# 'Chef IQ' —— **Summer 2026-08-10 定:Chef IQ 线就算 Tier 1**。
#
# US 项目没有 Weekly Tracker 行 → currentStatus / risk / crd / PO 一律留空
# (卡片不显示状态徽章和风险点,modal 相应字段显示 —)。Chefman 的 Status 行
# 是美方自由文本(如 'MP by middle of August for Costco CA followed by
# Amazon'),**原文进 modal 的 US Status 区,不提炼成阶段徽章** —— 提炼等于
# 替美方把阶段说具体(如 'waiting for first prototype' 到底算不算 Prototype
# 阶段,只有他们能定)。卡片改用蓝色 'US PD' 来源徽章,与中国卡区分。
# -------------------------------------------------------------

def _find_us_pd():
    """Newest readable 'US PD Update*.xlsx' in MONTHLY_DIR (same drop-in
    convention as _find_latest_pd_updates: next month's file needs no code edit)."""
    candidates = sorted(MONTHLY_DIR.glob('US PD Update*.xlsx'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    for p in candidates:
        try:
            ox.load_workbook(p, data_only=True)
            return p
        except Exception:
            continue
    return None


# 每个 sheet 一套 field → row 映射。'uf' 是列表(Chefman 有 4 个 Unique
# Feature 行,Chef IQ 只有 3 个);超出 uf1/uf2/uf3 的并入 uf3。
_US_SHEETS = [
    {
        'sheet': 'Chef IQ',
        'cols': ['C', 'D', 'E', 'F', 'G', 'H', 'I'],
        'rows': {
            'name': 2, 'pm': 3, 'tier': 4, 'market': 5, 'factory': 6,
            'sample': 7, 'tooled': 8, 'image': 9, 'brand': 10, 'model': 11,
            'desc': 12, 'msrp': 13, 'cost': 14, 'buffer': 16, 'port': 17,
            'duty': 18, 'hc40': 19, 'tooling': 20, 'insp': 21,
            'top': 22, 'uf': [23, 24, 25], 'comp': 26, 'rjdiff': 27,
            'notes': [28, 29, 30],
        },
        # row 2 标签写的是 'Category' 但填的是产品名 → 品类按列硬映射。
        # 3 台 iQ Mini Oven 归 Oven(台面烤箱,与中国 Oven 区同类)、
        # 3 个 iQ Sense 归 Thermometer(与中国 C60 同区)、Smart Kettle 归 Kettle。
        'cat_by_col': {
            'C': 'Oven', 'D': 'Oven', 'E': 'Oven',
            'F': 'Thermometer', 'G': 'Kettle',
            'H': 'Thermometer', 'I': 'Thermometer',
        },
        # Tier 列源值 'Chef IQ' → Tier 1 (Summer 2026-08-10)
        'tier_override': '1',
        'title_with_name': True,   # model 有重复(3 个项目都是 CQ60),标题带产品名
    },
    {
        'sheet': 'Chefman',
        'cols': ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
                 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
                 'Z', 'AA', 'AB', 'AC', 'AD', 'AE'],
        'rows': {
            'cat': 1, 'name': 2, 'pm': 3, 'tier': 4, 'market': 5, 'factory': 6,
            'sample': 7, 'tooled': 8, 'status': 9, 'image': 10, 'model': 11,
            'desc': 12, 'cost': 13, 'buffer': 14, 'port': 15, 'hc40': 16,
            'tooling': 17, 'insp': 18,
            'top': 19, 'uf': [20, 21, 22, 23], 'comp': 24,
            'notes': [25, 26, 27],
        },
        'cat_by_col': None,        # row 1 是真品类
        'tier_override': None,
        'title_with_name': False,  # model 唯一,标题用 model(与中国卡同形)
    },
]

# 美方品类写法 → 报表品类。只列 normalize_category() 认不出或会另开小分区的,
# 其余(Air Fryer / Kettle / Coffee Maker / Slow cooker / Microwave / Thermometer
# / Water dispenser / Slushie / Deep fryer / Blender / Rice Cooker / Grill)
# 走 normalize_category() 自动归一。子串匹配,小写比对。
_US_CAT_OVERRIDE = [
    ('toaf', 'Oven'),                    # Toaster Oven Air Fryer = 台面烤箱
    ('c38 with steam', 'Air Fryers'),
    ('double basket airfryer', 'Air Fryers'),   # 'airfryer' 连写,自动规则匹配不到
    ('everything maker', 'Everything Maker'),   # C58 圆/方/mini 三个,自成一区
    ('vacuum sealer', 'Vaccum Sealer'),  # 跟随中国 PD Table 现有拼法,保证同区
    ('espresso', 'Coffee'),
]

# PM 写法归一。同一份文件里 Mayer / Mayer Rosen 混用;Jennifer (CSM) 与中国侧
# Jenifer 是同一人(CSM 杭州)。不归一筛选器会裂成两个选项。
# 其余名字原样保留,不猜。
_US_PM_NORM = {
    'mayer rosen': 'Mayer',
    'jennifer (csm)': 'Jenifer',
    'jennifer': 'Jenifer',
}

_US_BULLET_RE = re.compile(r'^[•·●▪\-\*]+[ \t]*')


def _us_thumb(img_bytes):
    """raw image bytes → base64 JPEG data URI (same thumb size/quality as the
    China images so US and CN cards render identically). '' on failure."""
    try:
        from PIL import Image as PILImage
    except ImportError:
        return ''
    try:
        pil = PILImage.open(io.BytesIO(img_bytes))
        if pil.mode in ('RGBA', 'LA', 'P'):
            pil = pil.convert('RGBA')
            bg = PILImage.new('RGB', pil.size, (255, 255, 255))
            bg.paste(pil, mask=pil.split()[-1])
            pil = bg
        elif pil.mode != 'RGB':
            pil = pil.convert('RGB')
        pil.thumbnail((IMAGE_THUMB_SIZE, IMAGE_THUMB_SIZE), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil.save(buf, format='JPEG', quality=IMAGE_JPEG_QUALITY, optimize=True)
        return 'data:image/jpeg;base64,' + base64.b64encode(buf.getvalue()).decode('ascii')
    except Exception:
        return ''


def _us_clean(v):
    """cellstr + '#VALUE!' (broken refs / image cells) → ''."""
    s = cellstr(v).replace('\xa0', ' ').strip()
    return '' if s == '#VALUE!' else s


def _us_lines(s, strip_bullets=False):
    """Drop blank lines; optionally strip per-line bullet chars. Newlines kept —
    template renders detail values with white-space:pre-line."""
    lines = [ln.strip() for ln in str(s).split('\n')]
    lines = [ln for ln in lines if ln]
    if strip_bullets:
        lines = [_US_BULLET_RE.sub('', ln).strip() for ln in lines]
        lines = [ln for ln in lines if ln]
    return '\n'.join(lines)


def _us_category(raw_cat):
    """US category text → report category bucket."""
    s = (raw_cat or '').strip()
    if not s:
        return ''
    sl = s.lower()
    for needle, target in _US_CAT_OVERRIDE:
        if needle in sl:
            return target
    return normalize_category(s)


def _us_pm(raw_pm):
    p = (raw_pm or '').strip()
    return _US_PM_NORM.get(p.lower(), p)


def load_us_pd():
    """Parse both US sheets into page1-item dicts (build_page1_data schema).

    Returns (items, dup_report) where dup_report lists US models that collide
    with a China PD Table SKU — printed for Summer to arbitrate (we never
    silently merge or drop: the two sides carry different info, US = commercial
    + US status, CN = production status from the Weekly Tracker).
    """
    path = _find_us_pd()
    if path is None:
        print('      US PD: no "US PD Update*.xlsx" found — skipped')
        return [], []
    wb = ox.load_workbook(path, data_only=True)

    # Images: both sheets store them as image-in-cell rich data.
    raw_imgs = _extract_image_in_cell_raw(path)
    over_imgs = {}   # traditional floating images, keyed (sheet, row) → bytes
    for sname in wb.sheetnames:
        for im in getattr(wb[sname], '_images', []) or []:
            try:
                anc = im.anchor._from
                over_imgs.setdefault((sname, anc.row + 1, anc.col + 1), im._data())
            except Exception:
                continue

    items = []
    img_hits = 0
    for spec in _US_SHEETS:
        sname = spec['sheet']
        if sname not in wb.sheetnames:
            print(f'      US PD: sheet "{sname}" missing — skipped')
            continue
        ws = wb[sname]
        R = spec['rows']
        sheet_imgs = raw_imgs.get(sname, {})

        for col in spec['cols']:
            cidx = ox.utils.column_index_from_string(col)
            g = lambda key: _us_clean(ws[f'{col}{R[key]}'].value)

            model = g('model')
            name = g('name')
            raw_cat = g('cat') if 'cat' in R else ''
            # Chef IQ 的产品名在 'name' 行;Chefman 的产品名行是空的,品类行
            # (Rice Cooker / Pizza Maker ...) 才是人看的名字。
            display_name = name or raw_cat
            if not (model or display_name):
                continue        # 整列空

            if spec['title_with_name']:
                sku = (f'{display_name} ({model})'
                       if model and model.upper() != 'TBD' else display_name)
            else:
                sku = model if model and model.upper() != 'TBD' else f'{display_name} (model TBD)'
            sku = normalize_sku(sku)

            # ---- Tier: Chef IQ 线整体算 Tier 1;Chefman 读源值(1 / 1.5 / 2)
            tier = spec['tier_override'] or g('tier')
            if tier.endswith('.0'):
                tier = tier[:-2]

            # ---- 数值格式化
            msrp = g('msrp') if 'msrp' in R else ''
            if msrp:
                try:
                    msrp = f'${float(msrp):,.2f}'
                except ValueError:
                    pass
            duty = g('duty') if 'duty' in R else ''
            if duty:
                try:
                    duty = f'{float(duty) * 100:.1f}%'
                except ValueError:
                    pass
            cost = _us_lines(g('cost'))
            if cost and cost[0].isdigit():      # 裸数字成本(48.68)补 $
                cost = '$' + cost

            # ---- Sales sample ETA: 两个子行(7=ETA, 8=办公室已有的开模样)
            # Chef IQ 的 C/D/E 列 7:8 是合并单元格(一个答案盖两行) → row8 读到
            # 的是同一个值,去重。
            sample = g('sample')
            tooled = g('tooled')
            if tooled and tooled != sample:
                sample = (f'{sample}\nTooled sample: {tooled}' if sample
                          else f'Tooled sample: {tooled}')

            # ---- Unique features: 超出 3 个的并入 uf3(page1 schema 只有 3 位)
            ufs = [_us_lines(g_, strip_bullets=True)
                   for g_ in (_us_clean(ws[f'{col}{r}'].value) for r in R['uf'])]
            ufs = [u for u in ufs if u]
            uf1 = ufs[0] if len(ufs) > 0 else ''
            uf2 = ufs[1] if len(ufs) > 1 else ''
            uf3 = '\n'.join(ufs[2:]) if len(ufs) > 2 else ''

            # ---- Notes: Tooling kickoff 是 US 表独有字段,中国 schema 没有对应
            # 位置 → 拼进 note1 前部(modal Notes 区可见),不塞进 Est. Inspection
            # 免得串义。
            notes = [_us_clean(ws[f'{col}{r}'].value) for r in R['notes']]
            notes = [n for n in notes if n]
            tooling = g('tooling')
            if tooling:
                notes.insert(0, f'Tooling: {tooling}')
            note1 = notes[0] if notes else ''
            note2 = '\n'.join(notes[1:]) if len(notes) > 1 else ''

            # ---- 图片
            image = ''
            irow = R.get('image')
            if irow:
                b = sheet_imgs.get((irow, cidx)) or over_imgs.get((sname, irow, cidx))
                if b:
                    image = _us_thumb(b)
                    if image:
                        img_hits += 1

            items.append({
                'sku': sku,
                'category': _us_category(raw_cat) if raw_cat else spec['cat_by_col'].get(col, ''),
                'tier': tier,
                'brand': brand_for_sku(model, g('brand') if 'brand' in R else
                                       ('Chef iQ' if sname == 'Chef IQ' else 'Chefman')),
                'description': g('desc'),
                'topFeature': _us_lines(g('top'), strip_bullets=True),
                'uf1': uf1, 'uf2': uf2, 'uf3': uf3,
                'msrp': msrp,
                'sampleETA': sample,
                'poPlaced': '', 'poStatus': '', 'poBuyer': '',
                'estInspection': g('insp'),
                'factory': g('factory'),
                'market': g('market'),
                'cost': cost,
                'buffer': g('buffer'),
                'port': g('port'),
                'duty': duty,
                'hc40': g('hc40'),
                'compModel': g('comp'),
                'rjDiff': _us_lines(g('rjdiff')) if 'rjdiff' in R else '',
                'note1': note1,
                'note2': note2,
                'pmSection': f'US — {sname}',
                # 无 Weekly Tracker 行 → 状态/风险/CRD 一律空(见段头注释)
                'currentStatus': '', 'risk': '', 'crd': '', 'pv': '',
                'pm': _us_pm(g('pm')),
                'image': image,
                # US 专属:来源标记 + 美方 Status 原文
                'source': 'US',
                'usSheet': sname,
                'usStatus': _us_lines(g('status')) if 'status' in R else '',
                'usModel': model,
                'usProduct': display_name,
            })

    per_sheet = []
    for spec in _US_SHEETS:
        n = sum(1 for i in items if i['usSheet'] == spec['sheet'])
        per_sheet.append(f'{spec["sheet"]}={n}')
    print(f'      US PD source: {path.name}')
    print(f'      US PD: {len(items)} projects ({", ".join(per_sheet)}), '
          f'{img_hits} with images')
    return items, []


def report_us_overlap(us_items, pd_main, tracker_rows):
    """列出 US 项目与中国侧同号 / 近号的清单(供 Summer 裁决是否合并)。
    只报告,不改数据 —— 两边信息不同源(US=美方商务+状态,CN=Tracker 生产状态),
    合并谁压谁是业务裁决。"""
    cn_skus = set(pd_main.keys()) | {r['sku'] for r in tracker_rows if r.get('sku')}

    def base(s):
        return re.sub(r'[^A-Z0-9]', '', str(s or '').upper())

    cn_by_base = {}
    for s in cn_skus:
        cn_by_base.setdefault(base(s), set()).add(s)

    exact, near = [], []
    for it in us_items:
        m = it.get('usModel') or ''
        first = m.split('/')[0].strip()
        b = base(first)
        if not b:
            continue
        if b in cn_by_base:
            exact.append((it['sku'], sorted(cn_by_base[b])))
            continue
        hits = sorted({c for bb, cs in cn_by_base.items() for c in cs
                       if bb and len(b) >= 3 and (bb.startswith(b) or b.startswith(bb))})
        if hits:
            near.append((it['sku'], hits))
    if exact:
        print(f'      !! US/CN 同号 {len(exact)} 组(两张卡并存,待 Summer 裁决合并):')
        for us, cn in exact:
            print(f'         US {us}  ==  CN {", ".join(cn)}')
    if near:
        print(f'      ~  US/CN 近号 {len(near)} 组(疑似同项目不同版本):')
        for us, cn in near:
            print(f'         US {us}  ~~  CN {", ".join(cn)}')
    return exact, near


def drop_cn_pm_projects(us_items, tracker_rows, pd_main):
    """US region 里剔掉挂在中国 PM 名下的项目(Summer 2026-08-10 定)。

    美方那张表把一部分中国线项目也列了进来(C55 / C56 / C62 / C60 等),报送人
    写的是中国 PM 本人。这些项目在中国视图里已经有卡,留在 US region 属于重复,
    且会让 US 的 PM 筛选器混进中国 PM 名字。

    判定口径:US 行的 Project Manager 与 Weekly Tracker 的 PM 完全同名(已过
    PM 归一,见 _US_PM_ALIASES)。'Emma/Dan' 这类复合值不算同名,保留。
    中国侧没有对应 SKU 的会单独警示 —— 摘掉后它在报表里就完全看不到了。
    """
    cn_pms = {r['pm'] for r in tracker_rows if r.get('pm')}
    if not cn_pms:
        return us_items

    cn_skus = set(pd_main.keys()) | {r['sku'] for r in tracker_rows if r.get('sku')}

    def base(s):
        return re.sub(r'[^A-Z0-9]', '', str(s or '').upper())

    cn_bases = {base(s) for s in cn_skus}

    kept, dropped = [], []
    for it in us_items:
        if it.get('pm') in cn_pms:
            dropped.append(it)
        else:
            kept.append(it)

    if dropped:
        print(f'      -- US region 剔除 {len(dropped)} 张(PM 与中国侧同名 → 属中国线项目):')
        orphans = []
        for it in dropped:
            b = base((it.get('usModel') or '').split('/')[0])
            has_cn = bool(b) and any(cb == b or cb.startswith(b) or b.startswith(cb)
                                     for cb in cn_bases if cb)
            mark = '中国侧有对应卡' if has_cn else '!! 中国侧无对应,摘掉后报表里看不到'
            print(f'         [{it.get("pm")}] {it["sku"]} — {mark}')
            if not has_cn:
                orphans.append(it['sku'])
        if orphans:
            print(f'      !! 待 Summer 裁决: {len(orphans)} 个只在美方表里出现的项目被一并摘掉: '
                  f'{", ".join(orphans)}')
    return kept


if __name__ == '__main__':
    main()
