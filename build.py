"""
build.py — China PD Monthly Report HTML builder

Reads three xlsx data sources, joins by SKU (exact match only — never fuzzy),
and renders the report HTML by substituting JSON into template.html.

Sources:
  1. Weekly Tracker/China_PD_Weekly_Tracker_WK{xx}.xlsx  (project progress, ground truth)
  2. Monthly PD Report/Summers_Monthly_PD_Table.xlsx     (commercial info)
  3. Monthly PD Report/Project list.xlsx — China Projects sheet  (Sales white-list)

Output:
  Monthly PD Report/China_PD_Monthly_Report_{Mon}{Year}.html

Rotation rule:
  Within the same month-filename, keep latest + previous (_prev suffix).
  Older versions are deleted on each run.
"""

import base64
import io
import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl as ox

# -------------------------------------------------------------
# Paths — derive from __file__ so the script is session-id independent.
# This file lives at <BASE>/Monthly PD Report/build.py, so BASE is two levels up.
# -------------------------------------------------------------
MONTHLY_DIR = Path(__file__).resolve().parent
BASE = MONTHLY_DIR.parent
WEEKLY_DIR = BASE / 'Weekly Tracker'
# Sandbox uploads dir (Cowork-injected). Fallback when OneDrive Files On-Demand
# corrupts the project copy. Check both for the freshest readable Tracker.
# Derive from $HOME (Cowork sets HOME=/sessions/<id>) so it follows the current
# session automatically; the prior hardcoded session id breaks on every reboot.
_HOME = os.environ.get('HOME', '')
if _HOME and '/sessions/' in _HOME:
    UPLOADS_DIR = Path(_HOME) / 'mnt' / 'uploads'
else:
    UPLOADS_DIR = Path('/nonexistent-uploads')


def _safe_exists(p):
    """Path.exists() that tolerates PermissionError on cross-session mounts."""
    try:
        return p.exists()
    except (PermissionError, OSError):
        return False


def _find_latest_tracker():
    """Highest WKn xlsx that openpyxl can actually open. Falls back to uploads
    if the OneDrive copy has a corrupt zip footer."""
    candidates = []
    for d in (WEEKLY_DIR, UPLOADS_DIR):
        if _safe_exists(d):
            candidates.extend(d.glob('China_PD_Weekly_Tracker_WK*.xlsx'))
            candidates.extend(d.glob('China PD Weekly Tracker WK*.xlsx'))
    candidates = [p for p in candidates if 'backup' not in p.name.lower()]
    if not candidates:
        return WEEKLY_DIR / 'China_PD_Weekly_Tracker_WK17.xlsx'  # legacy default
    def wk_num(p):
        import re as _re
        m = _re.search(r'WK(\d+)', p.name)
        return int(m.group(1)) if m else 0
    candidates.sort(key=lambda p: (wk_num(p), p.stat().st_mtime), reverse=True)
    for p in candidates:
        try:
            ox.load_workbook(p, data_only=True)
            return p
        except Exception:
            continue
    return candidates[0]  # let main() raise the real error


TRACKER_PATH = _find_latest_tracker()
PDTABLE_PATH = MONTHLY_DIR / 'Summers_Monthly_PD_Table.xlsx'
PROJLIST_PATH = MONTHLY_DIR / 'Project list.xlsx'
# Source of embedded product images. Auto-detect the latest 'China PD updates *.xlsx'
# in MONTHLY_DIR by mtime, so Summer can drop in next month's file (e.g. 'China PD
# updates May 2026.xlsx') without editing build.py.
def _find_latest_pd_updates():
    """Newest readable 'China PD updates *.xlsx'. Project first, then uploads
    (OneDrive Files On-Demand bug)."""
    candidates = []
    for d in (MONTHLY_DIR, UPLOADS_DIR):
        if _safe_exists(d):
            candidates.extend(d.glob('China PD updates *.xlsx'))
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for p in candidates:
        try:
            ox.load_workbook(p, data_only=True)
            return p
        except Exception:
            continue
    return candidates[0] if candidates else None

PDUPDATES_PATH = _find_latest_pd_updates()
TEMPLATE_PATH = MONTHLY_DIR / 'template.html'
TRANSLATIONS_PATH = MONTHLY_DIR / 'translations.json'

# Output naming: month abbreviation + 4-digit year
NOW = datetime.now()
MONTH_NAME = NOW.strftime('%b')   # 'Apr'
YEAR = NOW.strftime('%Y')          # '2026'
# Override for current build (data is April 2026)
MONTH_NAME = 'Apr'
YEAR = '2026'

OUT_NAME = f'China_PD_Monthly_Report_{MONTH_NAME}{YEAR}.html'
OUT_PATH = MONTHLY_DIR / OUT_NAME
PREV_PATH = MONTHLY_DIR / f'China_PD_Monthly_Report_{MONTH_NAME}{YEAR}_prev.html'

# English-translated output (for US Sales)
OUT_NAME_EN = f'China_PD_Monthly_Report_{MONTH_NAME}{YEAR}_EN.html'
OUT_PATH_EN = MONTHLY_DIR / OUT_NAME_EN
PREV_PATH_EN = MONTHLY_DIR / f'China_PD_Monthly_Report_{MONTH_NAME}{YEAR}_EN_prev.html'

# Scratchpad for safe-write (avoids OneDrive zip-truncation issue).
# Use the env var so it follows the current session automatically.
SCRATCH = Path(os.environ.get('CLAUDE_SCRATCH', '/tmp/pd_report_scratch'))
SCRATCH.mkdir(parents=True, exist_ok=True)

# -------------------------------------------------------------
# Constants
# -------------------------------------------------------------
PIPELINE_LABELS = [
    'Kick off', 'Detail Design', 'Prototype', 'Tooling',
    'FOT', 'EB', 'Culinary EB', 'Culinary Claims',
    'PP', 'Culinary PP', 'MP',
]
# Inspection merged into MP (Summer 2026-05-04: 'inspection 和 MP 是一样的')

# Map currentStatus value → pipeline label index
STATUS_TO_PIPELINE = {
    'Kick off': 0, 'Kick Off': 0, 'kick off': 0,
    'Detail Design': 1, 'Design': 1,
    'Prototype': 2,
    'Tooling': 3, 'Tooling Launch': 3,
    'FOT': 4,
    'EB': 5, 'EB1': 5, 'EB2': 5,
    'Culinary EB': 6,
    'Culinary Claims': 7,
    'PP': 8,
    'Culinary PP': 9,
    'MP': 10, 'MP中': 10, 'Inspection': 10,
}

# Umbrella SKU expansion for Page 1 (Sales-facing cards).
# Some PD Table SKUs are 'umbrella' entries that bundle several color/finish
# variants under one row. Page 2 (Pipeline) and Page 3 (Tracker) keep them
# merged because dev progress is shared, but Sales wants to see each visible
# variant as its own card. For each umbrella SKU we list the variant SKUs that
# should render as separate cards — each variant copies all commercial info
# from the umbrella row but uses its own image (or shares the umbrella's image
# when the column has only one).
# Confirmed by Summer 2026-04-30. ASK before adding more entries.
SPLIT_UMBRELLA_SKUS = {
    'RJ50-SFDAF-25D':     ['RJ50-SFDAF-25D(SS)', 'RJ50-BFDAF-25D(BLK)'],
    'RJ62-20A-Series':    ['RJ62-BLK', 'RJ62-WHT'],
    'RJ64-10-new colors': ['RJ64-10-PTC', 'RJ64-10-BTR', 'RJ64-10-LVD', 'RJ64-10-Aqu'],
}


# PM section ordering (HTML displays in this order)
PM_SECTION_ORDER = [
    'Cottee Wei — 空气炸锅 + T1 项目',
    'Rowling Luo — 烤箱 / 面包机 / 饭煲 / 慢炖锅 / 油炸锅',
    'Serena Sun — ICEMAN / 咖啡 / 冰淇淋',
    'Chris Zhou — 烤盘 / 搅拌类 + MX 项目',
    'Liz Liu — 水壶 + 微波炉',
]

# -------------------------------------------------------------
# Helpers
# -------------------------------------------------------------
def cellstr(v):
    """Stringify a cell value, treating None as empty."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def normalize_sku(raw):
    """SKU normalization: drop newline annotations, strip whitespace.

    Examples:
      'RJ38-6T-lava grey\\n⚠️待确认项目' → 'RJ38-6T-lava grey'
      'RJ34-10C-D ' → 'RJ34-10C-D'

    NEVER drops semantic suffixes (SS/CA/MX/etc.) — exact match only.
    """
    if not raw:
        return ''
    s = str(raw)
    # Cut at first newline (annotation block)
    s = s.split('\n')[0]
    return s.strip()


def is_pm_section_header(value):
    """True if cell value is one of the 5 PM section headers."""
    if not value:
        return False
    s = str(value).strip()
    return s in PM_SECTION_ORDER


def clean_status(s):
    """Normalize Current Status string to a single-line value."""
    if not s:
        return ''
    return str(s).replace('\n', ' ').strip()


# -------------------------------------------------------------
# Source loaders
# -------------------------------------------------------------
def load_tracker(path):
    """Load Tracker. Returns list of dicts in row-order, plus pm_section list.

    Each dict has fields:
      num, sku, sku_raw, category, risk, pm, tier, last_update, current_status,
      issue, next_action, po_status, crd, stages (dict of stage_label → date_or_check),
      pm_section
    """
    if not path.exists():
        raise FileNotFoundError(f'Tracker not found: {path}')

    wb = ox.load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    rows = []
    current_section = ''

    # Tracker stage cols: C13-C24
    stage_label_map = {
        13: 'Kick off', 14: 'Detail Design', 15: 'Prototype',
        16: 'Tooling', 17: 'FOT', 18: 'EB',
        19: 'Culinary EB', 20: 'Culinary Claims',
        21: 'PP', 22: 'Culinary PP', 23: 'MP', 24: 'Inspection',
    }

    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c3 = ws.cell(r, 3).value  # SKU
        c1s = cellstr(c1)
        c3s = cellstr(c3)

        # PM section header
        if is_pm_section_header(c1s):
            current_section = c1s
            continue

        # Skip empty/non-data rows
        if not c3s:
            continue

        sku = normalize_sku(c3s)
        if not sku:
            continue

        # Stage cells: '✓' = completed past stage; date = scheduled/done; empty = not yet
        stages = {}
        for col, label in stage_label_map.items():
            v = ws.cell(r, col).value
            stages[label] = cellstr(v)

        rows.append({
            'num': cellstr(c1),
            'sku': sku,
            'sku_raw': c3s,
            'category': cellstr(ws.cell(r, 2).value),
            'risk': cellstr(ws.cell(r, 4).value),
            'pm': cellstr(ws.cell(r, 5).value),
            'tier': cellstr(ws.cell(r, 6).value),
            'last_update': cellstr(ws.cell(r, 7).value),
            'current_status': clean_status(ws.cell(r, 8).value),
            'issue': cellstr(ws.cell(r, 9).value),
            'next_action': cellstr(ws.cell(r, 10).value),
            'po_status': cellstr(ws.cell(r, 11).value),
            'crd': cellstr(ws.cell(r, 12).value),
            'stages': stages,
            'pm_section': current_section,
        })
    return rows


def load_pd_table(path):
    """Load Summers Monthly PD Table.

    Returns:
      main_skus: dict {sku → fields}  (rows above '▼ Gap Analysis' marker)
      pending_skus: list of (sku, category, pm_section) for rows in待确认 / Gap sections
    """
    if not path.exists():
        raise FileNotFoundError(f'PD Table not found: {path}')

    wb = ox.load_workbook(path, data_only=True)
    ws = wb['Product Info']

    main_skus = {}
    pending = []
    current_section = ''
    in_pending = False  # flips True at '▼ Gap Analysis' header

    # 24 column mapping (R1 header)
    col_map = {
        1: 'sku', 2: 'category', 3: 'tier', 4: 'brand',
        5: 'description', 6: 'top_feature',
        7: 'uf1', 8: 'uf2', 9: 'uf3',
        10: 'msrp', 11: 'sample_eta', 12: 'po_placed',
        13: 'est_inspection', 14: 'factory', 15: 'market',
        16: 'cost', 17: 'buffer', 18: 'port',
        19: 'duty', 20: 'hc40',
        21: 'comp_model', 22: 'rj_diff',
        23: 'note1', 24: 'note2',
    }

    for r in range(2, ws.max_row + 1):
        c1 = cellstr(ws.cell(r, 1).value)
        c2 = cellstr(ws.cell(r, 2).value)

        # Detect section transitions
        if c1.startswith('▼') or c1.startswith('⚠'):
            in_pending = True
            # Pending sub-section header (e.g., '▼ Cottee — 需补充商业信息')
            continue

        if is_pm_section_header(c1):
            current_section = c1
            in_pending = False
            continue

        if not c1:
            continue

        sku = normalize_sku(c1)
        if not sku:
            continue

        if in_pending:
            pending.append({
                'sku': sku,
                'sku_raw': c1,
                'category': c2,
                'pm_section': current_section,
            })
            continue

        # Main SKU row
        record = {'pm_section': current_section, 'sku_raw': c1}
        for col, key in col_map.items():
            v = ws.cell(r, col).value
            record[key] = cellstr(v)
        # Override sku with normalized
        record['sku'] = sku
        main_skus[sku] = record

    return main_skus, pending


def load_project_list(path):
    """Load Project list, China Projects sheet only. Returns set of SKUs (white-list)."""
    if not path.exists():
        raise FileNotFoundError(f'Project list not found: {path}')

    wb = ox.load_workbook(path, data_only=True)
    ws = wb['China Projects']

    skus = set()
    # Header at R7, data from R8. SKU column is C4 ('Model').
    for r in range(8, ws.max_row + 1):
        v = ws.cell(r, 4).value
        if not v:
            continue
        s = str(v)
        # Cells may have multi-line SKUs
        for piece in s.split('\n'):
            sku = normalize_sku(piece)
            if not sku:
                continue
            # Filter obviously non-SKU strings
            if sku.startswith(('RJ', 'C5', 'C6', 'C4', 'BF', 'GR')):
                skus.add(sku)
    return skus


# -------------------------------------------------------------
# Image extraction — pull embedded product renderings out of the PD updates
# xlsx and key them by SKU so we can render <img> tags in the HTML cards.
# -------------------------------------------------------------
# Tunables: thumbnail size (longest side, px) and JPEG quality. 300×~75% gives
# ~30–80 KB per image; 46 images ≈ 2–4 MB total embedded into the HTML.
IMAGE_THUMB_SIZE = 300
IMAGE_JPEG_QUALITY = 78


def _sku_image_aliases(sku):
    """Yield SKU aliases that should share the same product image.

    Confirmed by Summer 2026-04-30: a trailing parenthetical color/material
    code (SS = Stainless Steel, BLK = Black, WHT = White, etc.) names a
    visual variant of the same parent product. The parent SKU (without the
    parenthetical) can therefore reuse the same rendering image.

    Examples:
      'RJ50-SFDAF-25D(SS)'  → 'RJ50-SFDAF-25D(SS)', 'RJ50-SFDAF-25D'
      'RJ50-BFDAF-25D(BLK)' → 'RJ50-BFDAF-25D(BLK)', 'RJ50-BFDAF-25D'
      'RJ38-G4'             → 'RJ38-G4'

    NOTE: This is the ONLY image-aliasing rule. We never prefix-match,
    fuzzy-match, or strip non-parenthetical suffixes — those carry business
    meaning (CA / MX / V2 / etc.) and require explicit Summer approval.
    """
    if not sku:
        return
    yield sku
    m = re.match(r'^(.+?)\(\s*([A-Za-z][A-Za-z0-9]*)\s*\)\s*$', sku)
    if m:
        parent = m.group(1).strip()
        if parent and parent != sku:
            yield parent


def extract_sku_images(path):
    """Extract embedded product images from PD updates xlsx, keyed by SKU.

    Each sheet has products laid out horizontally — col B is the field label,
    cols C+ are one product per column. Row 8 holds the image, Row 10 holds
    the Model (SKU). When a single column carries multiple SKUs (color
    variants stacked in one cell, e.g. RJ64-10-PTC / BTR / LVD / Aqu) and the
    PM has placed several images side-by-side in the same anchor cell, we map
    them by reading order: images sorted top-to-bottom, left-to-right are
    assigned 1:1 to the SKUs in the cell. When there's only one image but
    multiple SKUs, all SKUs share the image. Multi-SKU cells with a single
    image keep the original 'all SKUs share' behaviour.

    Trailing '(SS)' / '(BLK)' / etc. parentheticals are color/material codes —
    we register the bare parent SKU as an additional alias so umbrella PD
    Table entries can match. See _sku_image_aliases().

    Returns: dict {sku → 'data:image/jpeg;base64,...'} for use as <img src=...>.
             Returns {} on failure (Pillow missing, file unreadable, etc.).
    """
    if not path or not path.exists():
        print(f'      WARN: PD updates file not found, skipping image extraction')
        return {}

    try:
        from PIL import Image as PILImage
    except ImportError:
        print(f'      WARN: Pillow not installed (pip install Pillow '
              '--break-system-packages), skipping image extraction')
        return {}

    images = {}
    skipped = 0
    try:
        wb = ox.load_workbook(path, data_only=True)
    except Exception as e:
        print(f'      WARN: could not open PD updates ({e}), skipping images')
        return {}

    # Bucket size for clustering rowOff values (EMU; 914400 = 1 inch). Images
    # within the same visual row have rowOff differences << 300000 EMU
    # (~0.33"), images in different visual rows differ by > 300000 EMU.
    ROW_BUCKET_EMU = 300000

    def _process(img_obj):
        """Process one openpyxl Image -> base64 JPEG data URI, or '' on error."""
        try:
            img_bytes = img_obj._data()
        except Exception:
            return ''
        try:
            pil = PILImage.open(io.BytesIO(img_bytes))
            if pil.mode in ('RGBA', 'LA'):
                bg = PILImage.new('RGB', pil.size, (255, 255, 255))
                bg.paste(pil, mask=pil.split()[-1])
                pil = bg
            elif pil.mode == 'P':
                pil = pil.convert('RGBA')
                bg = PILImage.new('RGB', pil.size, (255, 255, 255))
                bg.paste(pil, mask=pil.split()[-1])
                pil = bg
            elif pil.mode != 'RGB':
                pil = pil.convert('RGB')
            pil.thumbnail((IMAGE_THUMB_SIZE, IMAGE_THUMB_SIZE), PILImage.LANCZOS)
            buf = io.BytesIO()
            pil.save(buf, format='JPEG',
                     quality=IMAGE_JPEG_QUALITY, optimize=True)
            return ('data:image/jpeg;base64,'
                    + base64.b64encode(buf.getvalue()).decode('ascii'))
        except Exception:
            return ''

    def _register(sku, data_uri):
        for alias in _sku_image_aliases(normalize_sku(sku)):
            if alias and alias not in images:
                images[alias] = data_uri

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not getattr(ws, '_images', None):
            continue

        # Build column -> list-of-SKUs map for this sheet from Row 10.
        col_to_skus = {}
        for col_idx in range(3, ws.max_column + 1):  # C onwards
            v = ws.cell(10, col_idx).value
            if not v:
                continue
            skus_in_cell = []
            for line in str(v).split('\n'):
                line = line.strip()
                if not line:
                    continue
                # Strip annotations after the SKU: tab or 2+ spaces both signal
                # 'SKU<sep>colorname' (e.g. 'RJ64-10-PTC\tPistachio' or
                # 'RJ64-10-Aqu  Aqua'). Single space inside an SKU is rare so
                # we only split on tabs and runs of 2+ spaces.
                sku = re.split(r'\t|  +', line, maxsplit=1)[0].strip()
                if sku and any(c.isalnum() for c in sku):
                    skus_in_cell.append(sku)
            if skus_in_cell:
                col_to_skus[col_idx] = skus_in_cell

        # Group images by their target column (1-indexed).
        images_by_col = {}
        for img in ws._images:
            try:
                col = img.anchor._from.col + 1
                row_off = img.anchor._from.rowOff or 0
                col_off = img.anchor._from.colOff or 0
            except AttributeError:
                skipped += 1
                continue
            images_by_col.setdefault(col, []).append((row_off, col_off, img))

        for col, img_entries in images_by_col.items():
            # Find the SKU cell; tolerate 1-col drift.
            skus = col_to_skus.get(col)
            if not skus:
                for offset in (-1, 1):
                    skus = col_to_skus.get(col + offset)
                    if skus:
                        break
            if not skus:
                skipped += len(img_entries)
                continue

            # Sort by reading order: row bucket, then column offset.
            img_entries.sort(key=lambda t: (t[0] // ROW_BUCKET_EMU, t[1]))

            # When the column contains as many images as SKUs (e.g. ice cream
            # 4 colors, 4 images), assign 1:1 in reading order. Otherwise all
            # SKUs share each image (first one extracted wins per SKU).
            if len(img_entries) >= len(skus) > 1:
                for sku, (_r, _c, img) in zip(skus, img_entries):
                    data_uri = _process(img)
                    if not data_uri:
                        skipped += 1
                        continue
                    _register(sku, data_uri)
                # Any extra images beyond SKU count are ignored
            else:
                for _r, _c, img in img_entries:
                    data_uri = _process(img)
                    if not data_uri:
                        skipped += 1
                        continue
                    for sku in skus:
                        _register(sku, data_uri)

    msg = f'      images:     {len(images)} SKUs mapped'
    if skipped:
        msg += f' ({skipped} skipped)'
    print(msg)
    return images



# -------------------------------------------------------------
# Builders for the 4 JSON blocks
# -------------------------------------------------------------
def load_pd_config():
    """Load Monthly PD Report/pd_table_config.json. Used for ASI exclusion etc.
    Returns empty dict if file missing."""
    config_path = MONTHLY_DIR / 'pd_table_config.json'
    if not config_path.exists():
        return {}
    return json.loads(config_path.read_text(encoding='utf-8'))


def compute_mp_set(tracker_rows):
    """Return set of SKUs whose Weekly Tracker Current Status is MP or Inspection.
    Both stages count as 'Project Released' (Summer 2026-05-04: 'Inspection 和 MP
    是一样的')."""
    released_statuses = {'MP', 'INSPECTION'}
    return {r['sku'] for r in tracker_rows
            if (r.get('current_status') or '').strip().upper() in released_statuses}


def build_page1_data(pd_main, tracker_rows, white_list, asi_set, mp_set, images=None):
    """Page 1 = product cards. Driven by PD Table main rows.

    Joins each PD Table SKU with Tracker (for status/risk/crd/pm) by exact match.
    SKIPS SKUs that have no PD Table entry (no commercial info → can't draw card).
    SKIPS SKUs in `asi_set` (After Sales Improvement — Page 2/3 only) and `mp_set`
    (already MP — Project Released, separate stat card).

    `images` is an optional dict {sku → base64 data URI} produced by
    extract_sku_images(); SKUs without an image fall through to the placeholder
    icon in the template.
    """
    tracker_by_sku = {row['sku']: row for row in tracker_rows}
    if images is None:
        images = {}

    # Order: PM section order from PD Table, then row order within each section.
    items = []
    for section in PM_SECTION_ORDER:
        # Collect PD Table rows for this section, preserving insertion order
        section_skus = [(sku, rec) for sku, rec in pd_main.items() if rec['pm_section'] == section]
        for sku, rec in section_skus:
            tr = tracker_by_sku.get(sku)

            # Format cost: ensure $ prefix
            cost = rec.get('cost', '')
            if cost and not cost.startswith('$'):
                # Add $ if numeric-ish
                if any(c.isdigit() for c in cost):
                    cost = '$' + cost.lstrip('$ ')

            # Umbrella expansion: if this PD Table SKU is registered in
            # SPLIT_UMBRELLA_SKUS, render one card per listed variant. Each
            # variant card displays the variant's own SKU and image but copies
            # every commercial / Tracker field from the umbrella row.
            variants = SPLIT_UMBRELLA_SKUS.get(sku, [sku])
            umbrella_image = images.get(sku, '')

            for variant_sku in variants:
                # ASI / MP exclusion (Phase 2 — Project Released): these SKUs
                # are intentionally hidden from Page 1 cards. ASI lives in
                # config; MP is auto-detected from Tracker Current Status.
                if variant_sku in asi_set or variant_sku in mp_set:
                    continue
                # Variant image: prefer the variant's own; fall back to the
                # umbrella's image (one shared rendering for the whole group).
                variant_image = images.get(variant_sku, '') or umbrella_image

                # Page 1 uses PD Table as the authoritative source for tier and
                # category. If a SKU has empty tier/category in PD Table, it's
                # typically a US-side project that China is only tracking (not
                # commercializing) — those are intentionally hidden from the
                # Sales card view.
                item = {
                    'sku': variant_sku,
                    'category': rec.get('category', ''),
                    'tier': rec.get('tier', ''),
                    'brand': rec.get('brand', ''),
                    'description': rec.get('description', ''),
                    'topFeature': rec.get('top_feature', ''),
                    'uf1': rec.get('uf1', ''),
                    'uf2': rec.get('uf2', ''),
                    'uf3': rec.get('uf3', ''),
                    'msrp': rec.get('msrp', ''),
                    'sampleETA': rec.get('sample_eta', ''),
                    'poPlaced': rec.get('po_placed', ''),
                    'estInspection': rec.get('est_inspection', ''),
                    'factory': rec.get('factory', ''),
                    'market': rec.get('market', ''),
                    'cost': cost,
                    'buffer': rec.get('buffer', ''),
                    'port': rec.get('port', ''),
                    'duty': rec.get('duty', ''),
                    'hc40': rec.get('hc40', ''),
                    'compModel': rec.get('comp_model', ''),
                    'rjDiff': rec.get('rj_diff', ''),
                    'note1': rec.get('note1', ''),
                    'note2': rec.get('note2', ''),
                    'pmSection': section,
                    # Tracker-derived fields — variants share the umbrella's row
                    'currentStatus': tr['current_status'] if tr else '',
                    'risk': tr['risk'] if tr else '',
                    'crd': tr['crd'] if tr else '',
                    'pm': tr['pm'] if tr else '',
                    # White-list flag — check both umbrella and variant just in
                    # case Project List has either form
                    'onProjectList': (sku in white_list) or (variant_sku in white_list),
                    # Embedded base64 thumbnail (empty falls back to placeholder)
                    'image': variant_image,
                }
                items.append(item)
    return items


def build_page3_data(tracker_rows, asi_set=None):
    """Page 3 = Weekly Tracker rows. One per Tracker SKU.

    asi_set: optional set of SKUs flagged as After Sales Improvement; each row
    is tagged `isASI=true` so the front-end NPD/ASI filter can hide them.
    """
    asi_set = asi_set or set()
    items = []
    for i, row in enumerate(tracker_rows, 1):
        po_status, po_buyer = parse_po(row['po_status'])
        items.append({
            'num': str(i),
            'sku': row['sku'],
            'category': row['category'],
            'risk': row['risk'],
            'pm': row['pm'],
            'tier': row['tier'],
            'lastUpdate': row['last_update'],
            'currentStatus': row['current_status'],
            'issue': row['issue'],
            'nextAction': row['next_action'],
            'crd': row['crd'],
            'location': infer_location(row['next_action']),
            'pmSection': row['pm_section'],
            'stages': row['stages'],
            'poStatus': po_status,
            'poBuyer': po_buyer,
            'poRaw': row['po_status'],  # preserved for hover/tooltip if needed
            'isASI': row['sku'] in asi_set,
        })
    return items


def build_pipeline_data(tracker_rows, asi_set=None):
    """Page 2 = Pipeline Timeline. 12 stages, projects grouped by current stage.

    asi_set: optional set of SKUs flagged as After Sales Improvement; each
    project is tagged `isASI=true` so the front-end NPD/ASI filter can hide
    them and the on-page count badges can be recomputed on filter change.
    """
    asi_set = asi_set or set()
    counts = [0] * len(PIPELINE_LABELS)
    projects = [[] for _ in PIPELINE_LABELS]

    for row in tracker_rows:
        status = row['current_status']
        idx = STATUS_TO_PIPELINE.get(status)
        # Try case-insensitive / contains match for fallback
        if idx is None and status:
            sl = status.lower()
            for k, v in STATUS_TO_PIPELINE.items():
                if k.lower() == sl:
                    idx = v
                    break
        if idx is None:
            continue  # status doesn't map to a pipeline stage (e.g., '色样确认中', '—')
        counts[idx] += 1
        po_status, po_buyer = parse_po(row['po_status'])
        projects[idx].append({
            'sku': row['sku'],
            'category': row['category'],
            'pm': row['pm'],
            'risk': row['risk'],
            'status': PIPELINE_LABELS[idx],
            'action': row['next_action'],
            'poStatus': po_status,
            'poBuyer': po_buyer,
            'isASI': row['sku'] in asi_set,
        })

    return {'counts': counts, 'labels': PIPELINE_LABELS, 'projects': projects}


def build_summary_stats(page1, tracker_rows, asi_set, mp_set):
    """Auto-compute the 5 stats bar numbers.

    Rules (per Summer):
    - Total Projects: NPD + ASI active dev (only MP excluded). page1 is
      filtered to exclude both ASI and MP for cards, so add ASI-non-MP back.
    - High Risk / Medium Risk: page1 cards (NPD non-MP) — risk dimension
      doesn't apply to ASI for now.
    - Tier 1 (CSM): all T1 in Tracker including MP T1 (Summer's exception).
    - Project Released: total MP count (independent stat, replaces 'In MP').
    """
    visible = [p for p in page1 if p.get('category')]
    asi_non_mp = asi_set - mp_set  # ASI items not yet MP — count toward Total
    total = len(visible) + len(asi_non_mp)
    high = sum(1 for p in visible if p.get('risk') == '高')
    mid = sum(1 for p in visible if p.get('risk') == '中')
    t1 = sum(1 for r in tracker_rows if (r.get('tier') or '').strip() == '1')
    released = len(mp_set)
    return {'total': total, 'high': high, 'mid': mid, 't1': t1, 'released': released}


def build_released_data(tracker_rows, mp_set):
    """Data for the 'Project Released' stat card dropdown.
    Columns: SKU / PM / Category / PO info / CRD."""
    items = []
    for r in tracker_rows:
        if r['sku'] not in mp_set:
            continue
        po_status, po_buyer = parse_po(r.get('po_status', ''))
        items.append({
            'sku': r['sku'],
            'pm': r.get('pm', ''),
            'category': r.get('category', ''),
            'poStatus': po_status,
            'poBuyer': po_buyer,
            'poRaw': r.get('po_status', ''),
            'crd': r.get('crd', ''),
        })
    return items


# -------------------------------------------------------------
# Location heuristic — infer from Next Action keywords
# -------------------------------------------------------------
# US-side activities: Culinary review, design/artwork iterations, US-team confirmations
US_KEYWORDS = [
    'Culinary', 'culinary',
    'design', 'Design',
    'artwork', 'Artwork',
    'confirm', 'Confirm', '确认',  # confirm often means waiting on US sign-off
    'Andrew', 'Ryan',              # US team contacts
    'packaging', 'Packaging',
    'Sales', 'sales',
    'Pantone',
]
# China-side activities: prototyping, factory milestones, performance/life testing
CHINA_KEYWORDS = [
    '手板', '打样', '样品',
    'EB', 'PP', 'FOT',
    '寿命', '性能', '测试',
    '装配', '模具', '工厂',
    '大货', '色样', '注塑', '钣金',
    '量产', '试产',
]


def infer_location(next_action):
    """Infer China/US/Both/'' from Next Action text using keyword heuristics."""
    if not next_action:
        return ''
    has_us = any(kw in next_action for kw in US_KEYWORDS)
    has_cn = any(kw in next_action for kw in CHINA_KEYWORDS)
    if has_us and has_cn:
        return 'Both'
    if has_us:
        return 'US'
    if has_cn:
        return 'China'
    return ''


# -------------------------------------------------------------
# PO parsing — derive (status, buyer) from Tracker C11 free-text
# -------------------------------------------------------------
# Phrases that mean "no PO yet" / cancelled
PO_NEGATIVE_PHRASES = [
    '暂无订单', '无Open PO', '无open PO', '无PO', '无 PO',
    '项目Pending,无Open PO', '项目取消',
]
# Known buyer keywords. ORDER MATTERS — longer multi-word names listed first
# so "Canadian Tire" matches before "Canadian", "Walmart 3P" can be normalized
# to "Walmart" by listing "Walmart" alone (after multi-word variants).
PO_BUYER_KEYWORDS = [
    'Canadian Tire',
    "Sam's", 'Sams Club', 'Sams',
    "Kohl's",
    'PriceSmart',
    'Loblaws',
    'Menards',
    'Costco',
    'Walmart',
    'Amazon',
    'Target',
    'AAFES', 'Macy', 'BJ', 'AMZ',
    # Markets / channels (keep last so they don't shadow customer names)
    'MX', 'CA', 'EU', 'UK',
]


# -------------------------------------------------------------
# Translation helpers (Chinese → English for US Sales version)
# -------------------------------------------------------------
def load_translations():
    """Load Chinese→English dictionary. Returns empty dict if file missing."""
    if not TRANSLATIONS_PATH.exists():
        return {}
    import json as _json
    with open(TRANSLATIONS_PATH, 'r', encoding='utf-8') as f:
        return _json.load(f)


def translate(text, trans_dict):
    """Look up translation. If not found, return original (and we'll log it)."""
    if not text:
        return text
    return trans_dict.get(text, text)


# Risk display strings for English version (中文 risk values mapped to badges in JS,
# but for risk filter dropdown values and any raw display we map them here too).
RISK_ZH_TO_EN = {'高': 'High', '中': 'Medium', '低': 'Low', '—': '—'}


def translate_page1(items, trans_dict):
    """Translate page1 cards. Fields with potential Chinese: currentStatus, category, crd."""
    out = []
    for p in items:
        new_p = dict(p)
        for key in ['currentStatus', 'category', 'crd']:
            if p.get(key):
                new_p[key] = translate(p[key], trans_dict)
        out.append(new_p)
    return out


def translate_page3(items, trans_dict):
    """Translate page3 (Weekly Tracker) rows: issue / nextAction / currentStatus / category / poRaw / crd."""
    out = []
    for p in items:
        new_p = dict(p)
        for key in ['issue', 'nextAction', 'currentStatus', 'category', 'poRaw', 'crd']:
            if p.get(key):
                new_p[key] = translate(p[key], trans_dict)
        out.append(new_p)
    return out


def translate_pipeline(pipe, trans_dict):
    """Translate pipeline projects' action and category fields."""
    new_pipe = {
        'counts': pipe['counts'],
        'labels': pipe['labels'],
        'projects': [],
    }
    for stage_projs in pipe['projects']:
        new_stage = []
        for proj in stage_projs:
            np = dict(proj)
            if proj.get('action'):
                np['action'] = translate(proj['action'], trans_dict)
            if proj.get('category'):
                np['category'] = translate(proj['category'], trans_dict)
            new_stage.append(np)
        new_pipe['projects'].append(new_stage)
    return new_pipe


def report_untranslated(items_p1, items_p3, items_pipe, trans_dict):
    """Walk all data and report which Chinese strings have no translation."""
    import re as _re
    zh_pat = _re.compile(r'[一-鿿]')
    untrans = set()
    # page1
    for p in items_p1:
        for k in ['currentStatus', 'category', 'description', 'topFeature', 'crd']:
            v = p.get(k, '')
            if v and zh_pat.search(str(v)) and v not in trans_dict:
                untrans.add(v)
    # page3
    for p in items_p3:
        for k in ['issue', 'nextAction', 'currentStatus', 'category', 'poRaw', 'crd']:
            v = p.get(k, '')
            if v and zh_pat.search(str(v)) and v not in trans_dict:
                untrans.add(v)
    # pipeline
    for stage in items_pipe['projects']:
        for p in stage:
            v = p.get('action', '')
            if v and zh_pat.search(str(v)) and v not in trans_dict:
                untrans.add(v)
    return untrans


def parse_po(po_status_text):
    """Parse Tracker PO/订单状态 text → (status, buyer).

    Returns:
      status: 'YES' | 'NO' | ''   (placed / not placed / unknown)
      buyer:  string or ''        (extracted customer/channel name; '' if no PO or unknown)
    """
    if not po_status_text:
        return '', ''
    text = po_status_text.strip()
    if not text:
        return '', ''
    # Negative phrases → No PO
    if any(neg in text for neg in PO_NEGATIVE_PHRASES):
        return 'NO', ''
    # Try buyer extraction
    buyer = ''
    for kw in PO_BUYER_KEYWORDS:
        if kw in text:
            buyer = kw
            break
    # If we found a buyer keyword, it's a placed PO
    if buyer:
        return 'YES', buyer
    # Otherwise fallback: contains "for X" pattern → placed, buyer = X
    m = re.search(r'\bfor\s+([A-Z][A-Za-z\']+)', text)
    if m:
        return 'YES', m.group(1)
    # Has substantive text but no recognized pattern → still PO-placed (e.g., "Kohl's紧急订单")
    return 'YES', ''


# Map each PM section header → the English category names Sales would recognize.
# Banner shows these names instead of the messy raw category strings from PD Table.
PM_SECTION_TO_CATEGORIES = {
    'Cottee Wei — 空气炸锅 + T1 项目':
        ['Air Fryers'],
    'Rowling Luo — 烤箱 / 面包机 / 饭煲 / 慢炖锅 / 油炸锅':
        ['Ovens', 'Bread Maker', 'Rice Cooker', 'Slow Cooker', 'Deep Fryer'],
    'Serena Sun — ICEMAN / 咖啡 / 冰淇淋':
        ['Iceman', 'Coffee', 'Ice Cream'],
    'Chris Zhou — 烤盘 / 搅拌类 + MX 项目':
        ['Griddle', 'Blender'],
    'Liz Liu — 水壶 + 微波炉':
        ['Kettle', 'Microwave'],
}

# How many pending SKUs a PM must have before banner flags their categories.
# Singletons are usually one-off SKU-level issues, not a category-wide data gap.
BANNER_PM_THRESHOLD = 3


def build_banner_html(tracker_rows, pd_main, asi_set, mp_set):
    """Detect PMs with systemic data gaps and surface their domain categories.

    Logic (Phase 2 — adapted to new pure-mirror SOP):
      - For each Tracker SKU that is NOT in PD Table, NOT in ASI list, NOT MP:
        this is a gap — PM owes commercial info in next PD updates.
      - Count gaps per PM section. If ≥ BANNER_PM_THRESHOLD, flag categories.
      - Singletons (1-2 missing SKUs) are treated as SKU-level specifics and
        not surfaced — they're typically intentional gaps.

    Returns: HTML string for banner block (empty if no PM hits threshold).
    """
    pd_skus = set(pd_main.keys())
    pm_pending_count = {}
    for r in tracker_rows:
        sku = r.get('sku')
        if not sku:
            continue
        if sku in pd_skus or sku in asi_set or sku in mp_set:
            continue
        section = r.get('pm_section') or ''
        if section:
            pm_pending_count[section] = pm_pending_count.get(section, 0) + 1

    # Collect categories to flag
    flagged_cats = []
    flagged_pms = []
    for section in PM_SECTION_ORDER:
        if pm_pending_count.get(section, 0) >= BANNER_PM_THRESHOLD:
            cats = PM_SECTION_TO_CATEGORIES.get(section, [])
            flagged_cats.extend(cats)
            # Extract PM short name (first 2 words before em dash)
            pm_short = section.split('—')[0].strip()
            flagged_pms.append(pm_short)

    if not flagged_cats:
        return ''

    # Format category list: "A, B and C" English style
    if len(flagged_cats) == 1:
        cat_str = flagged_cats[0]
    elif len(flagged_cats) == 2:
        cat_str = f'{flagged_cats[0]} and {flagged_cats[1]}'
    else:
        cat_str = ', '.join(flagged_cats[:-1]) + f' and {flagged_cats[-1]}'

    banner = (
        f'<div class="data-banner">'
        f'<span class="banner-icon">⚠️</span>'
        f'<div><strong>Data note:</strong> '
        f'{cat_str} categor{"y" if len(flagged_cats) == 1 else "ies"} '
        f'currently lack complete commercial data — pending updates from related PM.</div>'
        f'</div>'
    )
    return banner


# -------------------------------------------------------------
# Render & rotate
# -------------------------------------------------------------
def render_template(template_text, page1, pipeline, page3, stats, banner, released):
    """Substitute placeholders with JSON / HTML."""
    out = template_text
    out = out.replace('{{PAGE1_DATA}}', json.dumps(page1, ensure_ascii=False))
    out = out.replace('{{PIPELINE_DATA}}', json.dumps(pipeline, ensure_ascii=False))
    out = out.replace('{{PAGE3_DATA}}', json.dumps(page3, ensure_ascii=False))
    out = out.replace('{{SUMMARY_STATS}}', json.dumps(stats, ensure_ascii=False))
    out = out.replace('{{RELEASED_DATA}}', json.dumps(released, ensure_ascii=False))
    out = out.replace('{{BANNER_BLOCK}}', banner)
    # Sanity: no placeholders should remain
    leftover = re.findall(r'\{\{[A-Z_]+\}\}', out)
    if leftover:
        raise RuntimeError(f'Unfilled placeholders: {leftover}')
    return out


def write_with_rotation(html_text, out_path, prev_path):
    """Apply rotation rule, then write new file safely."""
    if out_path.exists():
        try:
            shutil.move(str(out_path), str(prev_path))
            print(f'  rotation: {out_path.name} -> {prev_path.name}')
        except PermissionError:
            shutil.copyfile(str(out_path), str(prev_path))
            print(f'  rotation: copied {out_path.name} -> {prev_path.name}')

    scratch_file = SCRATCH / out_path.name
    scratch_file.write_text(html_text, encoding='utf-8')
    try:
        shutil.copyfile(str(scratch_file), str(out_path))
    except PermissionError:
        out_path.write_text(html_text, encoding='utf-8')
    print(f'  wrote: {out_path.name} ({len(html_text):,} chars)')


def main():
    print('=== China PD Monthly Report Builder ===')
    print(f'Output (CN): {OUT_NAME}')
    print(f'Output (EN): {OUT_NAME_EN}')
    print()

    print(f'[1/5] Loading template: {TEMPLATE_PATH.name}')
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f'Template missing: {TEMPLATE_PATH}')
    template = TEMPLATE_PATH.read_text(encoding='utf-8')

    print(f'[2/5] Loading data sources')
    print(f'      tracker:    {TRACKER_PATH.name}')
    tracker_rows = load_tracker(TRACKER_PATH)
    print(f'                  -> {len(tracker_rows)} SKU rows')

    print(f'      pd table:   {PDTABLE_PATH.name}')
    pd_main, pd_pending = load_pd_table(PDTABLE_PATH)
    print(f'                  -> {len(pd_main)} main SKUs, {len(pd_pending)} pending')

    print(f'      proj list:  {PROJLIST_PATH.name} (China Projects sheet)')
    white_list = load_project_list(PROJLIST_PATH)
    print(f'                  -> {len(white_list)} white-list SKUs')

    if PDUPDATES_PATH:
        print(f'      pd updates: {PDUPDATES_PATH.name} (extracting product images)')
    else:
        print(f'      pd updates: NONE FOUND (no embedded images this run)')
    images = extract_sku_images(PDUPDATES_PATH)

    print(f'[3/5] Building data blocks')
    config = load_pd_config()
    asi_set = set(config.get('after_sales_improvement', []))
    mp_set = compute_mp_set(tracker_rows)
    print(f'      ASI exclusion: {len(asi_set)} SKUs from config')
    print(f'      MP/Released set: {len(mp_set)} SKUs from Tracker Current Status="MP"')
    page1 = build_page1_data(pd_main, tracker_rows, white_list, asi_set, mp_set, images)
    print(f'      page1Data: {len(page1)} cards (ASI/MP excluded)')
    page3 = build_page3_data(tracker_rows, asi_set)
    print(f'      page3Data: {len(page3)} tracker rows')
    pipeline = build_pipeline_data(tracker_rows, asi_set)
    print(f'      pipelineData: counts={pipeline["counts"]} (total={sum(pipeline["counts"])})')
    stats = build_summary_stats(page1, tracker_rows, asi_set, mp_set)
    print(f'      summaryStats: {stats}')
    released = build_released_data(tracker_rows, mp_set)
    print(f'      releasedData: {len(released)} entries (Project Released dropdown)')

    banner = build_banner_html(tracker_rows, pd_main, asi_set, mp_set)
    if banner:
        print(f'[4/5] Banner ON')
    else:
        print(f'[4/5] Banner OFF')

    # 5a. Chinese version
    print(f'[5/5] Render + rotate (Chinese)')
    html_out = render_template(template, page1, pipeline, page3, stats, banner, released)
    write_with_rotation(html_out, OUT_PATH, PREV_PATH)

    # 5b. English version
    print(f'[5/5] Render + rotate (English)')
    trans = load_translations()
    print(f'      translations loaded: {len(trans)} entries')
    page1_en = translate_page1(page1, trans)
    page3_en = translate_page3(page3, trans)
    pipeline_en = translate_pipeline(pipeline, trans)

    untranslated = report_untranslated(page1_en, page3_en, pipeline_en, trans)
    if untranslated:
        print(f'      WARNING: {len(untranslated)} Chinese strings missing translation:')
        for s in sorted(untranslated)[:10]:
            print(f'        {s[:80]!r}')
        if len(untranslated) > 10:
            print(f'        ... +{len(untranslated)-10} more')
    else:
        print(f'      OK all Chinese strings translated')

    html_out_en = render_template(template, page1_en, pipeline_en, page3_en, stats, banner, released)
    write_with_rotation(html_out_en, OUT_PATH_EN, PREV_PATH_EN)
    print('Done.')


if __name__ == '__main__':
    main()
