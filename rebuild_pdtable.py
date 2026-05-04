"""
Rebuild Summer's Monthly PD Table from current China PD updates xlsx.

Pure mirror SOP:
  1. Read latest China PD updates → transpose 24 columns per sheet
  2. Inject manual_additions from config (e.g. PM hasn't added rows to PD updates yet)
  3. Output new Summer's Monthly PD Table (full overwrite, no merge with old)
  4. Auto-compare new PD Table SKUs vs latest Weekly Tracker SKUs
  5. Print diff grouped by PM (Summer broadcasts to PMs to align)

Config file: Monthly PD Report/pd_table_config.json
  - after_sales_improvement: SKUs that don't show on Page 1
  - umbrella_to_variants: umbrella SKU (Tracker) → variants (PD Table)
  - manual_additions: extra rows to inject (when PM hasn't put in PD updates yet)
  - sku_aliases (optional): Tracker SKU → canonical name
  - manually_excluded (optional): SKUs to drop entirely
"""
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# === Paths ===
# Derive from this script's location so the file can move sessions / containers
# without hardcoding session ids (Cowork rotates them every reboot).
import os as _os
PROJECT_DIR = Path(__file__).resolve().parent
TRACKER_DIR = PROJECT_DIR.parent / "Weekly Tracker"
_HOME = _os.environ.get("HOME", "")
if _HOME and "/sessions/" in _HOME:
    UPLOADS_DIR = Path(_HOME) / "mnt" / "uploads"
    SCRATCH_OUT = Path(_HOME) / "mnt" / "outputs" / "Summers_Monthly_PD_Table.xlsx"
else:
    UPLOADS_DIR = Path("/nonexistent-uploads")
    SCRATCH_OUT = Path("/tmp/Summers_Monthly_PD_Table.xlsx")
SCRATCH_OUT.parent.mkdir(parents=True, exist_ok=True)
FINAL_OUT = PROJECT_DIR / "Summers_Monthly_PD_Table.xlsx"
CONFIG_PATH = PROJECT_DIR / "pd_table_config.json"

def _safe_exists(p):
    try: return p.exists()
    except (PermissionError, OSError): return False

def find_pd_updates():
    """Find latest China PD updates xlsx. Try project folder first, fall back to uploads (OneDrive Files On-Demand bug workaround)."""
    candidates = []
    for d in (PROJECT_DIR, UPLOADS_DIR):
        candidates.extend(d.glob("China PD updates *.xlsx"))
    candidates = [p for p in candidates if _safe_exists(p)]
    if not candidates:
        raise FileNotFoundError("No China PD updates xlsx found in project or uploads")
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    # Try each in order; skip if BadZipFile
    for p in candidates:
        try:
            openpyxl.load_workbook(p, data_only=True)
            return p
        except Exception as e:
            print(f"  ⚠️  {p.name}: {e}, trying next")
    raise RuntimeError("All PD updates candidates are corrupt / unreadable")

def find_latest_tracker():
    """Find latest Weekly Tracker (highest WK number). Falls back to uploads if project copy is corrupt (OneDrive bug)."""
    candidates = []
    for d in (TRACKER_DIR, UPLOADS_DIR):
        candidates.extend(d.glob("China_PD_Weekly_Tracker_WK*.xlsx"))
        candidates.extend(d.glob("China PD Weekly Tracker WK*.xlsx"))
    candidates = [p for p in candidates if 'backup' not in p.name.lower()]
    if not candidates:
        raise FileNotFoundError("No Weekly Tracker xlsx found")
    def wk_num(p):
        m = re.search(r'WK(\d+)', p.name)
        return int(m.group(1)) if m else 0
    candidates.sort(key=lambda p: (wk_num(p), p.stat().st_mtime), reverse=True)
    for p in candidates:
        try:
            openpyxl.load_workbook(p, data_only=True)
            return p
        except Exception as e:
            print(f"  ⚠️  {p.name}: {e}, trying next")
    raise RuntimeError("All Tracker candidates are corrupt / unreadable")

# === Schema ===
COLUMNS = [
    ('SKU', 30), ('Category', 20), ('Tier', 6), ('Brand', 12), ('Description', 40),
    ('Top Feature', 35), ('Unique Feature 1', 25), ('Unique Feature 2', 25), ('Unique Feature 3', 25),
    ('MSRP', 10), ('Sales Sample ETA', 18), ('PO Placed?', 12), ('Est. 1st Inspection', 18),
    ('Factory', 15), ('Initial Market', 15), ('1st Cost Estimate', 15), ("Buffer Addt'l", 12),
    ('Port', 10), ('Duty', 10), ("40'HC", 10), ('Key Competitive Model', 25),
    ('Key RJ Brands Difference', 25), ('Note 1', 25), ('Note 2', 25),
]
COL_NAME_TO_IDX = {name: i+1 for i, (name, _) in enumerate(COLUMNS)}

LABEL_MAP = {
    'model': 1, 'category': 2, 'tier': 3, 'brand': 4, 'description': 5,
    'top feature': 6, 'msrp': 10, 'sales sample(s) eta': 11, 'po placed?': 12,
    'estimated 1st inspection': 13, 'factory': 14, 'initial market': 15,
    '1st cost estimate': 16, "buffer addt'l": 17, 'port': 18, 'duty (into us)': 19,
    "40'hc estimate": 20, 'key competitive model': 21, 'key rj brands difference': 22,
    'note (1)': 23, 'note (2)': 24,
}

PM_ORDER = [
    ('Cottee Wei', 'Cottee Wei — 空气炸锅 + T1 项目'),
    ('Rowling Luo', 'Rowling Luo — 烤箱 / 面包机 / 饭煲 / 慢炖锅 / 油炸锅'),
    ('Serena Sun', 'Serena Sun — ICEMAN / 咖啡 / 冰淇淋'),
    ('Chris Zhou', 'Chris Zhou — 烤盘 / 搅拌类 + MX 项目'),
    ('Liz Liu', 'Liz Liu — 水壶 + 微波炉'),
]

def normalize_pm(raw_pm, sheet_name):
    if raw_pm:
        s = str(raw_pm).strip()
        if s == 'Serena': return 'Serena Sun'
        if s == 'Tammy': return 'Chris Zhou'  # MX projects under Chris
        if s in ('Liz Liu', 'Cottee Wei', 'Serena Sun', 'Rowling Luo', 'Chris Zhou'):
            return s
        return s
    sheet_pm = {
        'Kettle': 'Liz Liu', 'Air Fryers': 'Cottee Wei', 'Microwaves': 'Liz Liu',
        'Coffee&Iceman': 'Serena Sun', 'Rice Cooker': 'Liz Liu', 'Juicer': 'Serena Sun',
        'OVEN&Bread maker&Deep fryer&Ric': 'Rowling Luo',
        'Roaster ovn&Waffle maker': 'Rowling Luo', 'Sourcing': 'Chris Zhou',
    }
    return sheet_pm.get(sheet_name, 'Chris Zhou')

SKU_RE = re.compile(r'^[A-Za-z]{1,5}\d+[\w\-()/]*$')

def parse_sku_cell(raw):
    if raw is None: return []
    s = str(raw).strip()
    if not s: return []
    skus = []
    for line in re.split(r'[\r\n]+', s):
        token = line.split('\t')[0].strip()
        if not token: continue
        if token.upper() in ('TBD', 'TBC'): continue
        if SKU_RE.match(token):
            skus.append(token)
    return skus

def fmt_value(label, val):
    if val is None: return None
    if isinstance(val, datetime): return val
    s = str(val).strip()
    if not s: return None
    if label == '1st cost estimate' and re.fullmatch(r'\d+(\.\d+)?', s):
        return f'${s}'
    return s

# === Loading ===
def load_config():
    if not CONFIG_PATH.exists():
        print(f"  ⚠️  No config found at {CONFIG_PATH}; using defaults (empty)")
        return {}
    with open(CONFIG_PATH, encoding='utf-8') as f:
        return json.load(f)

def load_pdupdates(src_path):
    wb = openpyxl.load_workbook(src_path, data_only=True)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_for_label = {}
        unique_feature_rows = []
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, 2).value
            if label is None: continue
            s = str(label).strip().lower()
            if s == 'unique feature':
                unique_feature_rows.append(r)
            elif s in LABEL_MAP:
                row_for_label[s] = r
        model_row = row_for_label.get('model')
        if not model_row: continue
        pm_row = None
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, 2).value
            if label and str(label).strip().lower() == 'project manager':
                pm_row = r; break
        for c in range(3, ws.max_column + 1):
            sku_cell = ws.cell(model_row, c).value
            skus = parse_sku_cell(sku_cell)
            if not skus: continue
            raw_pm = ws.cell(pm_row, c).value if pm_row else None
            pm = normalize_pm(raw_pm, sheet_name)
            data = {1: None}
            for label, col_idx in LABEL_MAP.items():
                if label == 'model': continue
                r = row_for_label.get(label)
                if r:
                    data[col_idx] = fmt_value(label, ws.cell(r, c).value)
            uf_vals = []
            for r in unique_feature_rows:
                v = fmt_value('unique feature', ws.cell(r, c).value)
                if v is not None: uf_vals.append(v)
            while len(uf_vals) < 3: uf_vals.append(None)
            data[7] = uf_vals[0]; data[8] = uf_vals[1]; data[9] = uf_vals[2]
            for sku in skus:
                row = dict(data)
                row[1] = sku
                row['_pm'] = pm
                row['_sheet'] = sheet_name
                records.append(row)
    return records

def apply_manual_additions(records, config):
    for entry in config.get('manual_additions', []):
        skus = entry.get('skus', [])
        pm = entry.get('pm', 'Chris Zhou')  # default fallback
        fields = entry.get('fields', {})
        for sku in skus:
            row = {'_pm': pm, '_sheet': '__manual__'}
            row[1] = sku
            for fname, val in fields.items():
                if fname in COL_NAME_TO_IDX:
                    row[COL_NAME_TO_IDX[fname]] = val
            for i in range(1, 25):
                if i not in row: row[i] = None
            records.append(row)
            print(f"  + manual add: {sku} (PM={pm})")
    return records

def apply_excluded(records, config):
    excluded = set(config.get('manually_excluded', []))
    if not excluded: return records
    before = len(records)
    records = [r for r in records if r.get(1) not in excluded]
    if len(records) < before:
        print(f"  - excluded {before - len(records)} SKU(s) per manually_excluded list")
    return records

# === Writing ===
def write_xlsx(records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Product Info'
    cg10 = Font(name='Century Gothic', size=10)
    cg10_bold = Font(name='Century Gothic', size=10, bold=True)
    cg10_white_bold = Font(name='Century Gothic', size=10, bold=True, color='FFFFFFFF')
    pm_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin = Side(border_style='thin', color='FFCCCCCC')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for ci, (name, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(1, ci, name)
        c.font = cg10_bold; c.alignment = align; c.border = border

    by_pm = {pm: [] for pm, _ in PM_ORDER}
    by_pm['__other__'] = []
    for rec in records:
        pm = rec['_pm']
        if pm in by_pm:
            by_pm[pm].append(rec)
        else:
            by_pm['__other__'].append(rec)

    cur_row = 2
    for pm_key, header_text in PM_ORDER:
        recs = by_pm[pm_key]
        if not recs: continue
        c = ws.cell(cur_row, 1, header_text)
        c.font = cg10_white_bold; c.fill = pm_fill; c.alignment = align
        for ci in range(2, len(COLUMNS) + 1):
            cc = ws.cell(cur_row, ci)
            cc.fill = pm_fill; cc.font = cg10_white_bold
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(COLUMNS))
        cur_row += 1
        for rec in recs:
            for ci in range(1, len(COLUMNS) + 1):
                v = rec.get(ci)
                cell = ws.cell(cur_row, ci, v)
                cell.font = cg10; cell.alignment = align; cell.border = border
                if ci in (11, 13) and isinstance(v, datetime):
                    cell.number_format = 'm/d/yyyy'
            cur_row += 1

    if by_pm['__other__']:
        orange = PatternFill(start_color='FFED7D31', end_color='FFED7D31', fill_type='solid')
        c = ws.cell(cur_row, 1, '⚠️ Other / 未归类 PM')
        c.font = cg10_white_bold; c.fill = orange; c.alignment = align
        for ci in range(2, len(COLUMNS) + 1):
            ws.cell(cur_row, ci).fill = orange
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=len(COLUMNS))
        cur_row += 1
        for rec in by_pm['__other__']:
            for ci in range(1, len(COLUMNS) + 1):
                v = rec.get(ci)
                cell = ws.cell(cur_row, ci, v)
                cell.font = cg10; cell.alignment = align; cell.border = border
            cur_row += 1

    for ci, (_, w) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'
    wb.save(SCRATCH_OUT)
    print(f"  ✓ Saved scratch → {SCRATCH_OUT}")
    try:
        shutil.copy(SCRATCH_OUT, FINAL_OUT)
        print(f"  ✓ Copied to OneDrive → {FINAL_OUT.name}")
    except PermissionError:
        FINAL_OUT.write_bytes(SCRATCH_OUT.read_bytes())
        print(f"  ✓ Direct-wrote to OneDrive (PermissionError fallback)")

# === Tracker comparison ===
def clean_sku(raw):
    if raw is None: return ''
    return str(raw).split('\n')[0].strip()

def load_pd_table_skus_with_pm():
    wb = openpyxl.load_workbook(FINAL_OUT, data_only=True)
    ws = wb.active
    out = {}
    cur_pm = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        cat = ws.cell(r, 2).value
        fill = ws.cell(r, 1).fill.start_color.rgb if ws.cell(r, 1).fill else None
        if fill == 'FF4472C4':
            txt = str(v) if v else ''
            if 'Cottee' in txt: cur_pm = 'Cottee Wei'
            elif 'Rowling' in txt: cur_pm = 'Rowling Luo'
            elif 'Serena' in txt: cur_pm = 'Serena Sun'
            elif 'Chris' in txt: cur_pm = 'Chris Zhou'
            elif 'Liz' in txt: cur_pm = 'Liz Liu'
            continue
        if not v: continue
        s = clean_sku(v)
        if s: out[s] = (cur_pm or '', str(cat).strip() if cat else '')
    return out

def load_tracker_skus(tracker_path, config):
    """Returns dict: SKU → (PM, category, current_status)."""
    wb = openpyxl.load_workbook(tracker_path, data_only=True)
    ws = wb.active
    aliases = config.get('sku_aliases', {})
    out = {}
    pm_full = {'Cottee': 'Cottee Wei', 'Rowling': 'Rowling Luo', 'Serena': 'Serena Sun',
               'Chris': 'Chris Zhou', 'Liz': 'Liz Liu'}
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, 3).value
        pm = ws.cell(r, 5).value
        cat = ws.cell(r, 2).value
        status = ws.cell(r, 8).value  # Current Status
        if not sku: continue
        s = clean_sku(sku)
        if not s: continue
        canonical = aliases.get(s, s)
        pm_clean = str(pm).strip() if pm else ''
        out[canonical] = (pm_full.get(pm_clean, pm_clean),
                          str(cat).split('\n')[0].strip() if cat else '',
                          str(status).strip() if status else '')
    return out

def compare_pd_vs_tracker(config):
    tracker_path = find_latest_tracker()
    print(f"\n  Reading Tracker: {tracker_path.name}")
    pd_skus = load_pd_table_skus_with_pm()
    tracker_skus = load_tracker_skus(tracker_path, config)
    asi = set(config.get('after_sales_improvement', []))
    umbrella = config.get('umbrella_to_variants', {})
    variant_to_umbrella = {v: u for u, vs in umbrella.items() for v in vs}

    # Tracker → PD: missing means PM hasn't added to PD updates yet
    # Filters: ASI excluded (not Page 1 anyway), MP excluded (released, no business info needed)
    missing_in_pd = []
    mp_in_tracker = []
    for sku, info in tracker_skus.items():
        status = info[2] if len(info) >= 3 else ''
        # MP first — Project Released counts MP or Inspection (Summer 5-04: same thing)
        if status.upper() in ('MP', 'INSPECTION'):
            mp_in_tracker.append((sku, info))
            continue  # released, no PD Table card needed
        if sku in asi: continue  # ASI (non-MP) not expected on PD Table
        if sku in pd_skus: continue
        # Umbrella check: if tracker has umbrella, PD has any variant
        if sku in umbrella:
            if any(v in pd_skus for v in umbrella[sku]):
                continue
        missing_in_pd.append((sku, info))

    # PD → Tracker: missing means rename / split / orphan
    missing_in_tracker = []
    for sku, info in pd_skus.items():
        if sku in tracker_skus: continue
        # Variant check: if PD variant, Tracker may have umbrella
        if sku in variant_to_umbrella:
            umb = variant_to_umbrella[sku]
            if umb in tracker_skus: continue
        missing_in_tracker.append((sku, info))

    return tracker_path.name, pd_skus, tracker_skus, missing_in_pd, missing_in_tracker, mp_in_tracker

def print_diff(tracker_name, pd_skus, tracker_skus, missing_in_pd, missing_in_tracker, mp_in_tracker):
    print(f"\n{'='*70}")
    print(f"PD Table vs {tracker_name} comparison")
    print(f"{'='*70}")
    print(f"PD Table: {len(pd_skus)} SKU   |   Tracker: {len(tracker_skus)} SKU   |   MP/Released: {len(mp_in_tracker)}")
    print(f"\n--- A) Tracker 有但 PD Table 没有 — {len(missing_in_pd)} 个 (已过滤 ASI 和 MP) ---")
    by_pm = {}
    for s, info in missing_in_pd:
        pm = info[0] or '(no PM)'
        by_pm.setdefault(pm, []).append((s, info[1]))
    for pm in sorted(by_pm.keys()):
        print(f"\n  【{pm}】 ({len(by_pm[pm])})")
        for s, c in by_pm[pm]:
            print(f"    {s:42} | {c}")
    print(f"\n--- B) PD Table 有但 Tracker 没有 — {len(missing_in_tracker)} 个 ---")
    by_pm2 = {}
    for s, info in missing_in_tracker:
        pm = info[0] or '(no PM)'
        by_pm2.setdefault(pm, []).append((s, info[1]))
    for pm in sorted(by_pm2.keys()):
        print(f"\n  【{pm}】 ({len(by_pm2[pm])})")
        for s, c in by_pm2[pm]:
            print(f"    {s:42} | {c}")
    print(f"\n--- C) Tracker 已 MP — {len(mp_in_tracker)} 个 (Project Released, 不需要 PD Table info) ---")
    by_pm3 = {}
    for s, info in mp_in_tracker:
        pm = info[0] or '(no PM)'
        by_pm3.setdefault(pm, []).append((s, info[1]))
    for pm in sorted(by_pm3.keys()):
        print(f"\n  【{pm}】 ({len(by_pm3[pm])})")
        for s, c in by_pm3[pm]:
            print(f"    {s:42} | {c}")
    print()

# === Main ===
if __name__ == '__main__':
    print("=== Summer's Monthly PD Table -- rebuild ===\n")
    config = load_config()
    print(f"  Config: {len(config.get('after_sales_improvement', []))} ASI, "
          f"{len(config.get('umbrella_to_variants', {}))} umbrella maps, "
          f"{len(config.get('manual_additions', []))} manual additions group(s)")
    src = find_pd_updates()
    print(f"  Reading PD updates: {src.name}")
    records = load_pdupdates(src)
    records = apply_excluded(records, config)
    records = apply_manual_additions(records, config)
    print(f"\n  Total records: {len(records)}")
    by_pm = {}
    for r in records:
        by_pm[r['_pm']] = by_pm.get(r['_pm'], 0) + 1
    for pm, n in sorted(by_pm.items()):
        print(f"    {pm}: {n}")
    write_xlsx(records)
    diff = compare_pd_vs_tracker(config)
    print_diff(*diff)
